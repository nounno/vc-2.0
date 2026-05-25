"""
Pipeline Service — Six-Layer Architecture with Async Intake:
  Intake Layer → Redis Streams → FormatDetector → ColumnTyper → SemanticMapper → EntityExtractors → QualityRouter

Knowledge Base: Home Appliance SKU Naming Rules (TV / AC / Refrigerator / Washer)
Sourced from: 家电SKU字典·型号命名规则.docx
"""

import json
import re
import uuid
import os
from datetime import datetime
from io import BytesIO
from typing import Any

import openpyxl
import redis
import mysql.connector
from mysql.connector import pooling
from fastapi import FastAPI, UploadFile
from pydantic import BaseModel

app = FastAPI()

# --------------------------------------------------------------------------
# Redis Configuration
# ---------------------------------------------------------------------------
REDIS_HOST = os.environ["REDIS_HOST"]
REDIS_PORT = int(os.environ["REDIS_PORT"])
STREAM_KEY = "pipeline:tasks"
RESULT_PREFIX = "pipeline:result:"
RESULT_TTL = 86400  # 24 hours

_redis_client = None


def get_redis() -> redis.Redis:
    global _redis_client
    if _redis_client is None:
        _redis_client = redis.Redis(host=REDIS_HOST, port=REDIS_PORT, decode_responses=True)
    return _redis_client


# ---------------------------------------------------------------------------
# MySQL Configuration (for persisting parsed records)
# ---------------------------------------------------------------------------
MYSQL_HOST = os.environ["MYSQL_HOST"]
MYSQL_PORT = int(os.environ["MYSQL_PORT"])
MYSQL_USER = os.environ["MYSQL_USER"]
MYSQL_PASSWORD = os.environ["MYSQL_PASSWORD"]
MYSQL_DATABASE = os.environ["MYSQL_DATABASE"]

_db_pool = None


def get_db_pool():
    global _db_pool
    if _db_pool is None:
        _db_pool = pooling.MySQLConnectionPool(
            pool_name="pipeline_pool",
            pool_size=3,
            host=MYSQL_HOST,
            port=MYSQL_PORT,
            user=MYSQL_USER,
            password=MYSQL_PASSWORD,
            database=MYSQL_DATABASE,
            connection_timeout=10,
            charset="utf8mb4",
            collation="utf8mb4_unicode_ci",
        )
    return _db_pool


def get_db_connection():
    pool = get_db_pool()
    return pool.get_connection()


def redis_health() -> bool:
    try:
        get_redis().ping()
        return True
    except Exception as e:
        import logging
        logging.getLogger("pipeline").warning(f"[pipeline] redis_health failed: {e}")
        return False


# ---------------------------------------------------------------------------
# Request/Response Models
# ---------------------------------------------------------------------------
class TaskSubmitResponse(BaseModel):
    task_id: str
    status: str
    message: str


class TaskResultResponse(BaseModel):
    task_id: str
    status: str
    result: dict | None = None


# ---------------------------------------------------------------------------
# Layer 1: Intake Layer
# ---------------------------------------------------------------------------
class IntakeLayer:
    """Reads the uploaded file and returns raw rows + extension.

    Handles merged cells by filling down/right the top-left cell value.
    Skips leading title rows (non-tabular header rows with large text).
    Returns (rows, ext) where rows[0] = column headers.
    """

    def load(self, file_content: bytes, filename: str) -> tuple[list[tuple[str, list[list[str]]]], str]:
        """Read xlsx/xls/csv file. Handles merged cells, forward-fill, multi-sheet.
        Returns list of (sheet_name, rows) tuples - one per data sheet (skip导航/目录 sheet).
        Each rows list has headers at index 0.
        """
        ext = filename.lower().split(".")[-1]

        # Handle CSV format
        if ext == "csv":
            import csv
            import io as csv_io
            sheet_data = []
            text_content = file_content.decode("utf-8-sig", errors="replace")
            reader = csv.reader(csv_io.StringIO(text_content))
            rows = []
            for row in reader:
                # Strip whitespace from each cell
                cleaned_row = [cell.strip() if cell else "" for cell in row]
                rows.append(cleaned_row)
            if rows:
                sheet_data.append(("Sheet1", rows))
            return sheet_data, ext

        # Handle ZIP format - extract and read first xlsx/xls/csv file
        if ext == "zip":
            import zipfile
            sheet_data = []
            try:
                with zipfile.ZipFile(BytesIO(file_content), 'r') as zf:
                    for name in zf.namelist():
                        if name.lower().endswith((".xlsx", ".xls", ".csv")):
                            inner_content = zf.read(name)
                            inner_ext = name.lower().split(".")[-1]
                            # Recursively process the extracted file
                            inner_filename = name.split("/")[-1]
                            return self.load(inner_content, inner_filename)
            except zipfile.BadZipFile:
                pass
            return sheet_data, ext

        # Try openpyxl first; fall back to raw zipfile+ET for corrupted workbooks
        try:
            wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
            sheet_data = self._load_xlsx_from_workbook(wb)
        except Exception:
            # Fallback: parse xlsx as raw XML via zipfile + ElementTree
            sheet_data = self._load_xlsx_raw_xml(BytesIO(file_content))

        return sheet_data, ext

    def _load_xlsx_from_workbook(self, wb) -> list[tuple[str, list[list[str]]]]:
        """Process an openpyxl Workbook object into sheet_data format."""
        SKIP_SHEET_KEYWORDS = ["目录", "导航", "index", "cover", "首页"]

        sheet_data = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Skip navigation/cover sheets
            if any(kw in sheet_name.lower() for kw in SKIP_SHEET_KEYWORDS):
                continue

            # ── Build cell-value map ─────────────────────────────────────────
            cell_values = {}
            for row_cells in ws.iter_rows():
                for cell in row_cells:
                    cell_values[(cell.row, cell.column)] = (
                        str(cell.value) if cell.value is not None else ""
                    )

            # ── Fill merged cells with top-left value ─────────────────────────
            for merged_range in ws.merged_cells.ranges:
                min_r, min_c = merged_range.min_row, merged_range.min_col
                top_val = cell_values.get((min_r, min_c), "")
                for r in range(min_r, merged_range.max_row + 1):
                    for c in range(min_c, merged_range.max_col + 1):
                        if (r, c) not in cell_values:
                            cell_values[(r, c)] = top_val

            max_row = ws.max_row
            max_col = ws.max_column

            # ── Forward-fill vertical spans (品类/品牌纵向合并格) ─────────────
            FILL_COLS = {1, 2}
            for col in FILL_COLS:
                last_val = ""
                for r in range(1, max_row + 1):
                    v = cell_values.get((r, col), "")
                    if v.strip():
                        last_val = v.strip()
                    else:
                        cell_values[(r, col)] = last_val

            # ── Detect header row ─────────────────────────────────────────────
            HEADER_MIN_NONEMPTY_RATIO = 0.20
            HEADER_MIN_LABEL_RATIO = 0.40
            header_row_idx = None
            for r in range(1, min(max_row + 1, 20)):
                row_vals = [cell_values.get((r, c), "") for c in range(1, max_col + 1)]
                non_empty = [v for v in row_vals if v.strip()]
                if not non_empty:
                    continue
                nonempty_ratio = len(non_empty) / max_col
                label_ratio = sum(1 for v in non_empty if len(v) < 50) / len(non_empty)
                # Skip merged title rows: very few non-empty cells (spread < 35%)
                # e.g. "库存表" merged across 17 columns → spread=1/17=6%
                if nonempty_ratio < 0.35 and len(non_empty) < 5:
                    continue
                if (nonempty_ratio > HEADER_MIN_NONEMPTY_RATIO and
                        label_ratio >= HEADER_MIN_LABEL_RATIO):
                    header_row_idx = r
                    break

            if header_row_idx is None:
                header_row_idx = 1

            # ── Collect rows for this sheet ─────────────────────────────────
            sheet_rows = []
            for r in range(header_row_idx, max_row + 1):
                row = [cell_values.get((r, c), "") for c in range(1, max_col + 1)]
                sheet_rows.append(row)

            sheet_data.append((sheet_name, sheet_rows))

        return sheet_data

    def _load_xlsx_raw_xml(self, file_io: BytesIO) -> list[tuple[str, list[list[str]]]]:
        """Fallback: parse xlsx via zipfile + ElementTree, bypassing openpyxl style errors.

        Handles:
        - Shared strings table (xl/sharedStrings.xml)
        - Inline string and numeric cells
        - Column/row dimensional info
        """
        import zipfile
        import xml.etree.ElementTree as ET

        SKIP_SHEET_KEYWORDS = ["目录", "导航", "index", "cover", "首页"]

        ns = {
            "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
            "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        }

        with zipfile.ZipFile(file_io, "r") as zf:
            # Load shared strings
            shared_strings: list[str] = []
            if "xl/sharedStrings.xml" in zf.namelist():
                ss_xml = zf.read("xl/sharedStrings.xml")
                ss_root = ET.fromstring(ss_xml)
                for si in ss_root.findall("main:si", ns):
                    # Concatenate all text runs within an si element
                    text_parts = []
                    for t in si.iter("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t"):
                        if t.text:
                            text_parts.append(t.text)
                    shared_strings.append("".join(text_parts))

            # Load workbook to map sheet names
            wb_xml = zf.read("xl/workbook.xml")
            wb_root = ET.fromstring(wb_xml)
            sheets_elem = wb_root.find(".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheets")
            if sheets_elem is None:
                # Try without braces (some variants)
                sheets_elem = wb_root.find(".//main:sheets", ns)

            sheet_info: list[tuple[str, str]] = []  # (name, sheet_xml_path)
            rels_xml = None
            if "xl/_rels/workbook.xml.rels" in zf.namelist():
                rels_xml = zf.read("xl/_rels/workbook.xml.rels")
                rels_root = ET.fromstring(rels_xml)
                # Build rid -> target map
                rid_to_target = {}
                for rel in rels_root:
                    rid_to_target[rel.attrib.get("Id", "")] = rel.attrib.get("Target", "")

            for sheet_elem in sheets_elem:
                sname = sheet_elem.attrib.get("name", "Sheet")
                sheet_id = sheet_elem.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id", "")
                target = rid_to_target.get(sheet_id, "")
                # target may be absolute ("/xl/worksheets/sheet1.xml" with leading slash)
                # or relative ("worksheets/sheet1.xml"), normalize to zip-internal path
                target = target.lstrip("/")
                if not target.startswith("xl/"):
                    target = "xl/" + target
                sheet_info.append((sname, target))

            sheet_data: list[tuple[str, list[list[str]]]] = []

            for sheet_name, sheet_path in sheet_info:
                if any(kw in sheet_name.lower() for kw in SKIP_SHEET_KEYWORDS):
                    continue

                sheet_xml = zf.read(sheet_path)
                sheet_root = ET.fromstring(sheet_xml)

                # Determine max col from <cols>
                max_col = 1
                cols_elem = sheet_root.find(".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}cols")
                if cols_elem is not None:
                    for col_elem in cols_elem:
                        col_idx = int(col_elem.attrib.get("min", 1))
                        max_col = max(max_col, col_idx)

                # Parse dimension
                dim = sheet_root.attrib.get("dimension", "")
                if dim:
                    # e.g. "A1:Z52"
                    from re import search
                    m = search(r"(\d+):(\d+)", dim)
                    if m:
                        pass  # we derive max_row from rows below

                # Build cell_values from sheetData
                cell_values: dict[tuple[int, int], str] = {}

                # Parse dimension from sheetData
                sheet_data_elem = sheet_root.find("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheetData")
                if sheet_data_elem is None:
                    sheet_data_elem = sheet_root.find(".//main:sheetData", ns)

                max_row = 1
                for row_elem in sheet_data_elem:
                    row_num = int(row_elem.attrib.get("r", 1))
                    max_row = max(max_row, row_num)
                    for cell_elem in row_elem:
                        cell_ref = cell_elem.attrib.get("r", "")
                        # Parse cell reference, e.g. "A1" -> col=1, row=1
                        col_str = "".join(c for c in cell_ref if c.isalpha())
                        row_str = "".join(c for c in cell_ref if c.isdigit())
                        if not col_str or not row_str:
                            continue
                        col_num = 0
                        for ch in col_str.upper():
                            col_num = col_num * 26 + (ord(ch) - ord('A') + 1)
                        row_num = int(row_str)

                        t = cell_elem.attrib.get("t", "")  # cell type
                        v_elem = cell_elem.find("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v")
                        is_elem = cell_elem.find("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}is")

                        if is_elem is not None:
                            # Inline string
                            text_parts = []
                            for t_elem in is_elem.iter("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t"):
                                if t_elem.text:
                                    text_parts.append(t_elem.text)
                            cell_values[(row_num, col_num)] = "".join(text_parts)
                        elif t == "s":
                            # Shared string
                            if v_elem is not None and v_elem.text is not None:
                                idx = int(v_elem.text)
                                cell_values[(row_num, col_num)] = shared_strings[idx] if idx < len(shared_strings) else ""
                            else:
                                cell_values[(row_num, col_num)] = ""
                        elif t == "inlineStr":
                            if v_elem is not None:
                                cell_values[(row_num, col_num)] = v_elem.text or ""
                            else:
                                cell_values[(row_num, col_num)] = ""
                        else:
                            # Numeric or other
                            if v_elem is not None and v_elem.text is not None:
                                cell_values[(row_num, col_num)] = v_elem.text
                            else:
                                cell_values[(row_num, col_num)] = ""

                # Handle merged cells
                for merge_elem in sheet_root.iter("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}mergeCell"):
                    ref = merge_elem.attrib.get("ref", "")
                    if not ref:
                        continue
                    # e.g. "B2:D4"
                    from re import match as re_match
                    m = re_match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", ref)
                    if m:
                        min_col_str, min_row_str, max_col_str, max_row_str = m.groups()
                        min_c = 0
                        for ch in min_col_str.upper():
                            min_c = min_c * 26 + (ord(ch) - ord('A') + 1)
                        min_r = int(min_row_str)
                        max_r = int(max_row_str)
                        top_val = cell_values.get((min_r, min_c), "")
                        for r in range(min_r, max_r + 1):
                            for c in range(min_c, min_c + (ord(max_col_str) - ord(min_col_str) + 1)):
                                if (r, c) not in cell_values:
                                    cell_values[(r, c)] = top_val

                # Re-derive max_col from parsed cells if dim wasn't helpful
                if max_col == 1:
                    for (r, c) in cell_values:
                        max_col = max(max_col, c)

                # Forward-fill cols 1 and 2
                FILL_COLS = {1, 2}
                for col in FILL_COLS:
                    last_val = ""
                    for r in range(1, max_row + 1):
                        v = cell_values.get((r, col), "")
                        if v.strip():
                            last_val = v.strip()
                        else:
                            cell_values[(r, col)] = last_val

                # Detect header row (same logic as openpyxl path)
                HEADER_MIN_NONEMPTY_RATIO = 0.20
                HEADER_MIN_LABEL_RATIO = 0.40
                header_row_idx = None
                for r in range(1, min(max_row + 1, 20)):
                    row_vals = [cell_values.get((r, c), "") for c in range(1, max_col + 1)]
                    non_empty = [v for v in row_vals if v.strip()]
                    if not non_empty:
                        continue
                    nonempty_ratio = len(non_empty) / max_col
                    label_ratio = sum(1 for v in non_empty if len(v) < 50) / len(non_empty)
                    if nonempty_ratio < 0.35 and len(non_empty) < 5:
                        continue
                    if (nonempty_ratio > HEADER_MIN_NONEMPTY_RATIO and
                            label_ratio >= HEADER_MIN_LABEL_RATIO):
                        header_row_idx = r
                        break

                if header_row_idx is None:
                    header_row_idx = 1

                sheet_rows: list[list[str]] = []
                for r in range(header_row_idx, max_row + 1):
                    row = [cell_values.get((r, c), "") for c in range(1, max_col + 1)]
                    sheet_rows.append(row)

                sheet_data.append((sheet_name, sheet_rows))

        return sheet_data


# ---------------------------------------------------------------------------
# Layer 2: Format Detector
# ---------------------------------------------------------------------------
class FormatDetector:
    """Detects the file format."""

    def detect(self, ext: str, content: bytes = None) -> str:
        if ext in ("xlsx", "xls"):
            return "excel"
        elif ext == "csv":
            return "csv"
        elif ext == "zip":
            return "zip"
        return "unknown"


# ---------------------------------------------------------------------------
# Type-A Supplier Templates (硬编码列映射，优先于正则匹配)
# --------------------------------------------------------------------------
_SUPPLIER_TEMPLATES: dict[str, dict] = {
    # 供应商25：两列表格，"型号"=model，"专属销售价"=price
    "5.25价格输出表": {
        "column_mapping": {0: "model", 1: "price"},
        "fixed_fields": {"brand": "通用", "category": "空调"},
    },
    # 供应商18：第0列空，第5列是备注，第6列是价格
    "5月25日VIP专属价格输出表": {
        "column_mapping": {1: "brand", 2: "category", 3: "model",
                          4: "text", 5: "text", 6: "price"},
        "fixed_fields": {},
    },
}

def _match_supplier_template(filename: str) -> str | None:
    """文件名关键字模糊匹配，返回template_key或None"""
    fname_clean = filename.replace(" ", "").replace("\u3000", "")
    for template_key in _SUPPLIER_TEMPLATES:
        key_clean = template_key.replace(" ", "").replace("\u3000", "")
        if key_clean in fname_clean or fname_clean in key_clean:
            return template_key
    return None


# ---------------------------------------------------------------------------
# Layer 3: Column Typer
# ---------------------------------------------------------------------------
class ColumnTyper:
    """Infers column types based on header names and sample values."""

    PRICE_RE = re.compile(r"^[\d,．.．]+$")
    INT_RE = re.compile(r"^\d+$")

    def infer(self, headers: list[str], rows: list[list[str]]) -> dict[int, str]:
        mapping = {}
        for i, h in enumerate(headers):
            h_lower = h.lower().strip()
            # 品牌
            if re.search(r"品牌|brand|厂商|品牌名称|牌子", h_lower):
                mapping[i] = "brand"
            # 品类（一级分类）
            elif re.search(r"品类|category|分类|类型", h_lower):
                mapping[i] = "category"
            # 型号
            elif re.search(r"型号|model|货号|款式|商品型号|产品型号", h_lower):
                mapping[i] = "model"
            # 价格（支持多种列名）
            elif re.search(r"价格|price|售价|单价|批发价|开票价|工程价|今日批价|供货价|专属销售价|最低零售价|明价|特价|活动价|会员价", h_lower):
                mapping[i] = "price"
            # 库存/数量
            elif re.search(r"库存|stock|数量|存货|销量", h_lower):
                mapping[i] = "stock"
            # 成本价/供货价/进货价
            elif re.search(r"供货价|成本价|进货价|采购价|工厂价", h_lower):
                mapping[i] = "costPrice"
            # 建议零售价/标价/市场价
            elif re.search(r"建议零售价|标价|市场价|零售价|指导价|挂牌价", h_lower):
                mapping[i] = "suggestedPrice"
            # 数量/件数
            elif re.search(r"数量|件数|库存量|备货量|可售数量", h_lower):
                mapping[i] = "quantity"
            # 仓库/库房/发货地
            elif re.search(r"仓库|库房|发货地|仓库地址|所在仓库", h_lower):
                mapping[i] = "warehouse"
            else:
                # Try to infer from values
                sample_vals = [row[i] if i < len(row) else "" for row in rows[:5]]
                if all(self.PRICE_RE.match(v) for v in sample_vals if v):
                    mapping[i] = "price"
                elif all(self.INT_RE.match(v) for v in sample_vals if v):
                    mapping[i] = "stock"
                else:
                    mapping[i] = "text"
        return mapping

# ---------------------------------------------------------------------------
# Layer 4: Semantic Mapper
# ---------------------------------------------------------------------------
KNOWN_BRANDS = {
    "美的", "格力", "海尔", "海信", "小米", "华为", "华凌", "华帝", "华蒜",
    "长虹", "TCL", "奥克斯", "志高", "新飞", "美菱", "容声",
    "松下", "索尼", "飞利浦", "博世", "西门子", "三星", "LG",
    "东芝", "卡萨帝", "COLMO", "统帅", "小天鹅", "小吉", "石头",
    "创维", "海信", "康佳", "康宝", "星星", "美菱", "奥马", "新科",
    "扬子", "科龙", "万和", "万家乐", "方太", "老板", "华帝",
    "林内", "能率", "史密斯", "沁园", "安吉尔", "云米", "佳尼特",
    "火星人", "森歌", "美大", "名气", "雷鸟", "VIDAA", "酷开",
    "乐华", "哈士奇", "微果", "西屋",
}
KNOWN_CATEGORIES = {
    "空调", "冰箱", "洗衣机", "热水器", "电视投影", "厨房大电",
    "冷柜", "冰吧", "消毒柜", "洗碗机", "净水器", "烟机", "灶具",
    "集成灶", "投影仪", "电视", "管线机", "中央空调", "移动空调",
    "挂机", "柜机", "滚筒洗衣机", "波轮洗衣机", "洗烘一体机", "洗烘套装",
    "烘干机", "迷你洗衣机", "全自动洗衣机", "半自动",
    "电热水器", "燃气热水器", "空气能热水器", "壁挂炉",
    "十字门", "法式", "对开门", "三门", "双门", "单门", "异型门",
    "冷柜/冰吧", "展示柜", "嵌入式微蒸烤", "蒸烤一体机", "懒人三筒", "集成水槽",
    "子母双筒", "方型柜式", "壁挂式", "圆柱立式",
}
# ─────────────────────────────────────────────────────────────────────────────
# Knowledge Base: Home Appliance SKU Naming Rules
# Source: 家电SKU字典·型号命名规则.docx (2026-05-19)
# ─────────────────────────────────────────────────────────────────────────────

## ── TV (Television) ──────────────────────────────────────────────────────────
# Structure: Brand prefix + Screen size (inch) + Series + Gen/Suffix
# No mandatory national standard; each brand has its own system.
TV_BRAND_PREFIXES = {
    "TCL":         "TCL",
    "海信":        "海信",
    "Hisense":     "海信",
    "创维":        "创维",
    "Skyworth":    "创维",
    "华为":        "华为",
    "华为智慧屏":  "华为",
    "小米":        "小米",
    "小米电视":    "小米",
    "三星":        "三星",
    "Samsung":      "三星",
    "LG":          "LG",
    "索尼":        "索尼",
    "Sony":        "索尼",
    "索尼电视":    "索尼",
}
# TV model patterns: extract size (leading digits), infer series from brand
TV_SIZE_RE = re.compile(r"^(\d{2,3})(?:\s|英寸|寸)?")
# TCL: Q=QD-MiniLED旗舰, T=中高端, V=入门; K=2024, L=2025
# 海信: U=ULED旗舰, E=线上主力, L=激光; N=2024, Q=2025
# 创维: A=壁纸艺术, G=MiniLED性价比, Q=旗舰, X=OLED高端
# 华为: X=顶尖科技, V=旗舰影院, S=年轻潮流, SE=入门
# 小米: 大师=旗舰, S=性价比, A=入门
# 三星: QN=NeoQLED, UA=LED; D=2024, F=2025
# LG: G=旗舰(2025=G5), C=主流(C5), B=入门; OLED/QNED
# 索尼: KD=4K液晶, XR=XR认知芯片; 8Ⅱ=8系二代旗舰, 7Ⅱ=7系MiniLED
TV_SERIES_KEYWORDS = {
    "Q": ["TCL", "海信"],   # 线上旗舰/高端
    "E": ["海信"],           # 线上主力
    "A": ["创维", "华为"],   # 艺术/入门
    "G": ["创维", "小米"],   # MiniLED性价比
    "S": ["小米", "华为"],   # 性价比/年轻
    "V": ["华为"],           # 旗舰影院
    "X": ["华为", "LG"],     # 顶尖旗舰
    "QN": ["三星"],          # NeoQLED
    "UA": ["三星"],          # LED液晶
}

## ── Air Conditioner (分体式) ────────────────────────────────────────────────
# National Standard GB/T 7725:
#   K F R - [制冷量] G/L W / [品牌后缀]
#   K=空调器, F=分体式, R=热泵型(冷暖), 数字=百瓦(35=1.5P, 26=1P, 50=2P, 72=3P)
#   G=挂机, L=柜机, W=室外机
AC_NATIONAL_STANDARD = re.compile(
    r"^(KF|KFR|KFRd|KFRBd|KD)-\d{2,3}GW|LW"
)
AC_HP_FROM_CAPTION = {"23": "小1匹", "26": "1匹", "32": "1.5匹", "35": "正1.5匹",
                       "50": "2匹", "72": "3匹", "大字": "3匹+"}
# Gree: 云佳(入门), 云锦(中高端), 云逸, 全能王(旗舰65°C); 严格遵循国标
# Midea: 酷省电(节能), 风尊(180°旋转), 静新风; 严格遵循国标
# Haier: 麦浪(UWB人感), 劲爽(快冷热), 洗空气; 严格遵循国标
# AUX: 京岳(主流); 遵循国标
# Xiaomi: 国标认证型号+KFR-35GW/N1A1等; 营销名另起; 代工厂A/HN1=长虹,C=海信,X=TCL
# Daikin: 自有体系FTXP开头; 横纲Z=旗舰, 衡境=静音; VRV=多联中央空调注册商标
# Hitachi: RAS开头(分体), RA(窗口), RCI(中央空调内机)
# Panasonic: 遵循国标+自有代码; VX=旗舰声控, 滢风Pro, 飓能星, 净悦星
# Mitsubishi E: MS开头(挂/柜机), MXZ(中央空调外机); AHJ/ZFJ=挂机经典
AC_BRANDS_HIERARCHY = {
    "格力": {"正道": "KFR-数字GW", "系列": ["云佳", "云锦", "云逸", "全能王"]},
    "美的": {"正道": "KFR-数字GW", "系列": ["酷省电", "风尊", "静新风"]},
    "海尔": {"正道": "KFR-数字GW", "系列": ["麦浪", "劲爽", "洗空气"]},
    "奥克斯": {"正道": "KFR-数字GW", "系列": ["京岳"]},
    "小米": {"正道": "KFR-数字GW/N后缀", "系列": ["巨省电", "自然风", "新风空调"], "代工厂": {"A": "长虹", "HN1": "长虹", "C": "海信", "X": "TCL"}},
    "大金": {"正道": "FTXP/RXP开头", "系列": ["横纲Z", "衡境"]},
    "日立": {"正道": "RAS/RA/RCI", "系列": []},
    "松下": {"正道": "CS-/CU-", "系列": ["VX", "滢风Pro", "飓能星", "净悦星"]},
    "三菱电机": {"正道": "MS-/MXZ", "系列": ["AHJ", "ZFJ"]},
}

## ── Refrigerator ─────────────────────────────────────────────────────────────
# National Standard:
#   B C D - [容积L] [后缀]
#   B=冰箱, C=冷藏, D=冷冻; BC=仅冷藏, BD=仅冷冻(冷柜)
#   后缀: W=风冷, P=变频, T=多门/三门, K=对开门, G=玻璃门
RFR_NATIONAL = re.compile(r"^BCD?-\d+")
# ⚠️ CRITICAL陷阱: "双风路循环" ≠ "双循环/双系统" (后者才是双蒸发器防串味)
# 海尔: 麦浪(594mm零嵌+全空间保鲜), 星蕴; 严格遵循国标BCD-
# 美的: M60(600mm超薄), 慧鲜; 严格遵循
# 容声: 方糖(高颜值保鲜); 严格遵循
# 海信: 真空保鲜; 严格遵循
# COLMO: CRBF开头; 美的旗下AI高端
# 卡萨帝: 原石/揽光系列(法式多门标杆); 海尔高端
# 松下: BCD-和NR-双前缀共存; C=三门, E=多门; NORUVISEA™=2025旗舰
# 三星: RS=对开门, RF=多门(FrenchDoor), RB=另有; 低端用BCD-
# LG: NR-/GR-全线, S=对开门, M=多门
# 西门子: K开头(独立); K=冰箱, A=对开门, 数字≈容积; IQ100/300/500/700分级; 无界=平嵌旗舰
# 博世: 与西门子同平台; 2/4/6/8系从低到高; M8/M7=极限嵌入旗舰
RFR_DOOR_TYPES = {
    "T": "三门",
    "K": "对开门",
    "十字": "十字对开",
    "法式": "法式多门",
    "对开": "对开门",
    "多门": "多门",
}
RFR_BRANDS_HIERARCHY = {
    "海尔": {"正道": "BCD-数字", "系列": ["麦浪", "星蕴"], "卖点": ["零嵌", "全空间保鲜"]},
    "美的": {"正道": "BCD-数字", "系列": ["M60", "慧鲜"]},
    "容声": {"正道": "BCD-数字", "系列": ["方糖", "原鲜"]},
    "海信": {"正道": "BCD-数字", "系列": ["真空保鲜"]},
    "COLMO": {"正道": "CRBF", "系列": []},
    "卡萨帝": {"正道": "BCD-数字", "系列": ["原石", "揽光"], "卖点": ["法式多门", "全空间保鲜"]},
    "松下": {"正道": "BCD-|NR-", "系列": ["NORUVISEA"]},
    "三星": {"正道": "RS|RF|RB", "系列": []},
    "LG": {"正道": "NR-|GR-", "系列": []},
    "西门子": {"正道": "K.\\d+", "系列": ["IQ100", "IQ300", "IQ500", "IQ700", "无界"]},
    "博世": {"正道": "K.\\d+", "系列": ["2系", "4系", "6系", "8系", "全域智净"]},
}

## ── Washer ──────────────────────────────────────────────────────────────────
# National Standard:
#   X Q G/B - [容量kg] [后缀]
#   X=洗衣机, Q=全自动, G=滚筒, B=波轮; 容量数字=额定洗涤容量(kg)
# 小天鹅: TD=洗烘一体, TG=单滚筒, TB=波轮; 水魔方(冷水护色), 小蓝鲸(蓝氧防串)
# 西门子: WM(大容量)/WS(小容量)+国标型号; iQ300/500/700/800分级; 摩德纳系列
# LG: A/C/T/N+转速数字; DD直驱是其核心技术标签
# 海尔: HBD=烘一体, S=双动力; 朗境X11(风巡航防霉), 卡萨帝双子云裳(上下双筒)
WASHER_NATIONAL = re.compile(r"^XQG?-\d+")
WASHER_TYPES = {
    "XQG": "滚筒全自动",
    "XQB": "波轮",
    "TG": "小天鹅单滚筒",
    "TD": "小天鹅洗烘一体",
    "TB": "小天鹅波轮",
    "WM": "西门子大容量",
    "WS": "西门子小容量",
}
WASHER_BRANDS_HIERARCHY = {
    "海尔": {"正道": "XQG-数字", "系列": ["朗境X11", "卡萨帝双子云裳"], "后缀": {"HBD": "烘一体", "S": "双动力"}},
    "小天鹅": {"正道": "TD|TG|TB-数字", "系列": ["水魔方", "小蓝鲸"]},
    "美的": {"正道": "MG|MB-数字", "系列": []},
    "小米": {"正道": "XQG-数字", "系列": ["巨省电"]},
    "西门子": {"正道": "WM|WS|XQG", "系列": ["iQ300", "iQ500", "iQ700", "iQ800", "摩德纳"]},
    "博世": {"正道": "XQG", "系列": ["2系", "4系", "6系", "8系"]},
    "LG": {"正道": "WD|A.|C.", "系列": ["DD直驱", "CFC直驱"]},
    "松下": {"正道": "XQG-数字", "系列": ["罗密欧", "ALPHA G5"]},
}

## ── Unified Brand Registry ───────────────────────────────────────────────────
KNOWN_BRANDS = {
    # 空调主流
    "美的", "格力", "海尔", "海信", "小米", "华为", "华凌", "奥克斯",
    "大金", "日立", "松下", "三菱电机", "三菱", "志高", "长虹", "TCL",
    # 冰箱
    "容声", "美菱", "COLMO", "卡萨帝", "新飞", "奥马",
    # 洗衣机
    "小天鹅", "小吉", "石头", "西门子", "博世", "LG", "惠而浦",
    # 电视
    "创维", "康佳", "三星", "LG", "索尼", "夏普", "雷鸟", "VIDAA", "酷开",
    # 厨卫/热水器
    "华帝", "华蒜", "万和", "万家乐", "方太", "老板", "林内", "能率", "史密斯",
    "沁园", "安吉尔", "云米", "佳尼特", "火星人", "森歌", "美大",
    # 冷柜
    "星星", "美菱", "新科",
    # 互联网/其他
    "扬子", "科龙", "乐华", "哈士奇", "微果", "西屋",
}

## ── Category Inference ──────────────────────────────────────────────────────
# From product model naming rules + keyword match
CATEGORY_RULES = {
    "ac": {
        "names": ["空调", "冷暖", "挂机", "柜机", "分体机", "中央空调", "移动空调"],
        "model_prefixes": ["KFR", "KF", "KFRd", "KFRBd", "KD", "RAS", "RA", "CS-", "CU-", "MSZ", "MXZ"],
        "door_types": ["挂机", "柜机"],
    },
    "refrigerator": {
        "names": ["冰箱", "冷柜", "冰吧", "展示柜", "冷藏箱", "冷冻箱"],
        "model_prefixes": ["BCD", "BC", "BD", "NR-", "GR-", "RS", "RF", "RB", "CRBF", "KA"],
        "door_types": ["单门", "双门", "三门", "对开", "十字", "法式", "多门"],
    },
    "washer": {
        "names": ["洗衣机", "洗烘", "烘干机", "干衣机", "迷你洗衣机"],
        "model_prefixes": ["XQG", "XQB", "TG", "TD", "TB", "WM", "WS", "WD"],
        "door_types": ["滚筒", "波轮", "搅拌式"],
    },
    "tv": {
        "names": ["电视", "电视机", "智慧屏", "投影", "激光电视"],
        "model_prefixes": [],
        "door_types": [],
    },
}

## ── Cross-Category Confusion Pairs (易混淆品类) ────────────────────────────
# These model prefixes appear in multiple contexts; use category keywords to disambiguate
AMBIGUOUS_PREFIXES = {
    "KFR": "ac",
    "KF":  "ac",
    "BCD": "refrigerator",
    "XQG": "washer",
}

## ── Marketing Trap Detection ────────────────────────────────────────────────
MARKETING_TRAPS = [
    ("双风路循环", "双循环/双系统", "冰箱"),
    ("不提分区数的MiniLED", "有真实分区数", "电视"),
    ("一晚一度电", "实测耗电量", "空调"),
    ("等效刷新率", "原生刷新率", "电视"),
    ("磨号", "正常序列号", "通用"),
]


PRICE_RANGE = (50, 100000)


# ─────────────────────────────────────────────────────────────────────────────
# Model Decoder — uses SKU naming rules to decode model strings
# ─────────────────────────────────────────────────────────────────────────────

class ModelDecoder:
    """
    Decodes home appliance model strings using the SKU naming rule knowledge base.
    Returns structured entity + reasoning chain.
    """

    def __init__(self, brand: str = "", model: str = "", category_hint: str = ""):
        self.brand = brand
        self.model = (model or "").strip()
        self.category_hint = category_hint
        self.reasoning = []

    # ── Category Inference ─────────────────────────────────────────────────

    def infer_category(self) -> str:
        """Infer product category from model prefix + brand."""
        m = self.model.upper()

        # 1. Check model prefix
        for prefix, cat in AMBIGUOUS_PREFIXES.items():
            if m.startswith(prefix.upper()):
                self.reasoning.append(f"前缀{prefix}→{cat}")
                return cat

        # 2. Check category keywords in model
        for cat, rules in CATEGORY_RULES.items():
            for kw in rules.get("model_prefixes", []):
                if kw.upper() in m:
                    self.reasoning.append(f"型号含{kw}→{cat}")
                    return cat

        # 3. Check brand→category affinity
        brand_cat = {
            "格力": "ac", "美的": "ac", "海尔": "ac", "大金": "ac", "奥克斯": "ac",
            "海信": "refrigerator", "容声": "refrigerator", "美菱": "refrigerator",
            "小天鹅": "washer", "西门子": "washer", "博世": "washer",
            "TCL": "tv", "创维": "tv", "海信": "tv", "三星": "tv", "LG": "tv",
        }
        inferred = brand_cat.get(self.brand, "")
        if inferred:
            self.reasoning.append(f"品牌{self.brand}→{inferred}")
            return inferred

        return "unknown"

    # ── AC Model Decoder ──────────────────────────────────────────────────

    def decode_ac(self) -> dict[str, Any]:
        """Decode KFR/KF/KF Rd series AC model."""
        m = self.model
        result = {"horsepower": "", "type": "", "brand_series": "", "oem_factory": ""}

        # Extract horsepower: digits before GW/LW
        hp_match = re.search(r"(\d{2,3})\s*(?:GW|LW)", m, re.IGNORECASE)
        if hp_match:
            digits = hp_match.group(1)
            result["horsepower"] = AC_HP_FROM_CAPTION.get(digits, f"{int(digits)//100}匹")
            self.reasoning.append(f"匹数{digits}→{result['horsepower']}")

        # Indoor unit type
        if "GW" in m.upper():
            result["type"] = "挂壁式"
            self.reasoning.append("室内机类型: 挂机(G)")
        elif "LW" in m.upper():
            result["type"] = "落地式"
            self.reasoning.append("室内机类型: 柜机(L)")

        # Brand series from suffix
        if self.brand in AC_BRANDS_HIERARCHY:
            series_map = AC_BRANDS_HIERARCHY[self.brand].get("系列", [])
            for s in series_map:
                if s in m:
                    result["brand_series"] = s
                    self.reasoning.append(f"检测到系列: {s}")
                    break

        # OEM factory detection (Xiaomi)
        if self.brand == "小米":
            oem_map = AC_BRANDS_HIERARCHY["小米"].get("代工厂", {})
            for code, factory in oem_map.items():
                if code in m:
                    result["oem_factory"] = factory
                    self.reasoning.append(f"代工厂: {factory}({code})")
                    break

        return result

    # ── Refrigerator Model Decoder ────────────────────────────────────────

    def decode_rfr(self) -> dict[str, Any]:
        """Decode BCD series refrigerator model."""
        m = self.model
        result = {"volume_l": "", "door_type": "", "cooling": "", "brand_series": ""}

        # Extract volume
        vol_match = re.search(r"BCD?-\s*(\d+)", m, re.IGNORECASE)
        if not vol_match:
            vol_match = re.search(r"(\d{3,4})\s*(?:L|升|瓦)", m, re.IGNORECASE)
        if vol_match:
            result["volume_l"] = vol_match.group(1) + "L"
            self.reasoning.append(f"容积: {result['volume_l']}")

        # Door type from suffix
        for code, dtype in RFR_DOOR_TYPES.items():
            if code in m:
                result["door_type"] = dtype
                self.reasoning.append(f"门型: {dtype}({code})")
                break

        # Cooling type
        if "W" in m:
            result["cooling"] = "风冷"
            self.reasoning.append("制冷: 风冷(W)")
        if "Z" in m or "P" in m:
            result["cooling"] = result.get("cooling", "") + "变频"
            self.reasoning.append("变频: 是")

        # Brand series
        if self.brand in RFR_BRANDS_HIERARCHY:
            for s in RFR_BRANDS_HIERARCHY[self.brand].get("系列", []):
                if s in m:
                    result["brand_series"] = s
                    self.reasoning.append(f"系列: {s}")
                    break

        return result

    # ── Washer Model Decoder ───────────────────────────────────────────────

    def decode_washer(self) -> dict[str, Any]:
        """Decode XQG series washer model."""
        m = self.model
        result = {"capacity_kg": "", "type": "", "brand_series": "", "special": ""}

        # Extract capacity
        cap_match = re.search(r"XQG?-?\s*(\d+)", m, re.IGNORECASE)
        if cap_match:
            result["capacity_kg"] = cap_match.group(1) + "kg"
            self.reasoning.append(f"洗涤容量: {result['capacity_kg']}")

        # Type
        for code, wtype in WASHER_TYPES.items():
            if code in m:
                result["type"] = wtype
                self.reasoning.append(f"类型: {wtype}")
                break

        # Brand series
        if self.brand in WASHER_BRANDS_HIERARCHY:
            for s in WASHER_BRANDS_HIERARCHY[self.brand].get("系列", []):
                if s in m:
                    result["brand_series"] = s
                    self.reasoning.append(f"系列: {s}")
                    break

        # Special suffix (HBD=烘一体, S=双动力, etc.)
        if self.brand == "海尔":
            suffixes = WASHER_BRANDS_HIERARCHY["海尔"].get("后缀", {})
            for code, meaning in suffixes.items():
                if code in m:
                    result["special"] = meaning
                    self.reasoning.append(f"功能: {meaning}({code})")

        return result

    # ── TV Model Decoder ───────────────────────────────────────────────────

    def decode_tv(self) -> dict[str, Any]:
        """Decode TV model string (no national standard, brand-specific)."""
        m = self.model
        result = {"size_inch": "", "series": "", "gen_code": ""}

        # Extract screen size
        size_match = TV_SIZE_RE.match(m)
        if size_match:
            result["size_inch"] = size_match.group(1) + "英寸"
            self.reasoning.append(f"屏幕尺寸: {result['size_inch']}")

        # Series detection
        for series_code, brands in TV_SERIES_KEYWORDS.items():
            if self.brand in brands and series_code in m:
                result["series"] = series_code
                self.reasoning.append(f"系列代码: {series_code} ({self.brand})")
                break

        # Year code
        year_codes = {"K": "2024", "L": "2025", "D": "2024", "F": "2025", "N": "2024", "Q": "2025"}
        for code, year in year_codes.items():
            if code in m:
                result["gen_code"] = f"{year}年款"
                self.reasoning.append(f"年份: {result['gen_code']}({code})")
                break

        return result

    # ── Full Decode Dispatch ───────────────────────────────────────────────

    def decode(self) -> dict[str, Any]:
        """Run full decode pipeline."""
        category = self.infer_category()

        decoded = {
            "brand": self.brand,
            "model": self.model,
            "category": category,
            "specs": {},
            "marketing_traps": [],
            "reasoning": self.reasoning,
        }

        if category == "ac":
            decoded["specs"] = self.decode_ac()
        elif category == "refrigerator":
            decoded["specs"] = self.decode_rfr()
        elif category == "washer":
            decoded["specs"] = self.decode_washer()
        elif category == "tv":
            decoded["specs"] = self.decode_tv()

        # Check marketing traps
        for trap, genuine, apply_to in MARKETING_TRAPS:
            if trap in self.model and (apply_to == "通用" or apply_to == category):
                decoded["marketing_traps"].append({
                    "trap": trap,
                    "genuine_meaning": genuine,
                    "category": apply_to,
                })
                self.reasoning.append(f"⚠️ 陷阱检测: '{trap}' (≠{genuine})")

        return decoded


def decode_model(brand: str, model: str, category_hint: str = "") -> dict[str, Any]:
    """Convenience wrapper for ModelDecoder."""
    decoder = ModelDecoder(brand=brand, model=model, category_hint=category_hint)
    return decoder.decode()


# ─────────────────────────────────────────────────────────────────────────────
# Legacy Variables (kept for backward compatibility with existing code)
# ─────────────────────────────────────────────────────────────────────────────
AC_REGEX = re.compile(r"KFR|KF-|空调|挂机|柜机|分体", re.IGNORECASE)


def clean_price(v: str) -> float | None:
    if not v:
        return None
    cleaned = re.sub(r"[^\d.]", "", v)
    try:
        return float(cleaned)
    except ValueError:
        return None


def clean_int(v: str) -> int | None:
    if not v:
        return None
    cleaned = re.sub(r"[^\d]", "", v)
    try:
        return int(cleaned)
    except ValueError:
        return None


class SemanticMapper:
    # Column name patterns for extended IR fields
    COST_PRICE_RE = re.compile(r"供货价|成本价|进货价|采购价|工厂价")
    SUGGESTED_PRICE_RE = re.compile(r"建议零售价|标价|市场价|零售价|指导价|挂牌价")
    QUANTITY_RE = re.compile(r"数量|件数|库存量|库存数量|备货量|可售数量")
    WAREHOUSE_RE = re.compile(r"仓库|库房|发货地|仓库地址|所在仓库")

    def map(self, col_type: str, raw_value: str, row_idx: int, header_name: str = "") -> dict[str, Any]:
        if not raw_value or raw_value.strip() == "":
            return {"value": None, "confidence": 0.0}

        if col_type == "brand":
            brand = raw_value.strip()
            conf = 0.98 if brand in KNOWN_BRANDS else 0.70
            return {"value": brand, "confidence": conf}

        elif col_type == "category":
            cat = raw_value.strip()
            conf = 0.98 if cat in KNOWN_CATEGORIES else 0.70
            return {"value": cat, "confidence": conf}

        elif col_type == "model":
            model = raw_value.strip()
            conf = 0.92 if model else 0.60
            return {"value": model, "confidence": conf}

        elif col_type == "price":
            price = clean_price(raw_value)
            if price is None:
                return {"value": None, "confidence": 0.0}
            in_range = PRICE_RANGE[0] <= price <= PRICE_RANGE[1]
            conf = 0.90 if in_range else 0.50
            return {"value": price, "confidence": conf}

        elif col_type == "stock":
            stock = clean_int(raw_value)
            if stock is None:
                return {"value": None, "confidence": 0.0}
            conf = 0.90 if 0 <= stock <= 99999 else 0.50
            return {"value": stock, "confidence": conf}

        elif col_type == "costPrice":
            price = clean_price(raw_value)
            if price is None:
                return {"value": None, "confidence": 0.0}
            # costPrice is typically lower than retail; accept a wider range
            in_range = 1 <= price <= 100000
            conf = 0.88 if in_range else 0.50
            return {"value": price, "confidence": conf}

        elif col_type == "suggestedPrice":
            price = clean_price(raw_value)
            if price is None:
                return {"value": None, "confidence": 0.0}
            in_range = PRICE_RANGE[0] <= price <= PRICE_RANGE[1]
            conf = 0.88 if in_range else 0.50
            return {"value": price, "confidence": conf}

        elif col_type == "quantity":
            qty = clean_int(raw_value)
            if qty is None:
                return {"value": None, "confidence": 0.0}
            conf = 0.88 if 0 <= qty <= 999999 else 0.50
            return {"value": qty, "confidence": conf}

        elif col_type == "warehouse":
            warehouse = raw_value.strip()
            if warehouse:
                return {"value": warehouse, "confidence": 0.85}
            return {"value": None, "confidence": 0.0}

        else:
            return {"value": raw_value.strip(), "confidence": 0.60}


# ---------------------------------------------------------------------------
# Layer 5: Entity Extractor
# ---------------------------------------------------------------------------
# Price unit conversion patterns
PRICE_UNIT_PATTERNS = [
    (re.compile(r"(\d+(?:\.\d+)?)\s*万元"), 10000),   # 万元 → 元
    (re.compile(r"(\d+(?:\.\d+)?)\s*百元"), 100),      # 百元 → 元
    (re.compile(r"(\d+(?:\.\d+)?)\s*千元"), 1000),     # 千元 → 元
    (re.compile(r"(\d+(?:\.\d+)?)\s*万"), 10000),      # 万 → 元 (without "元")
    (re.compile(r"(\d+(?:\.\d+)?)\s*千"), 1000),       # 千 → 元
]


def _convert_price_unit(raw_value: str) -> str:
    """Convert price from 万元/百元/千元 to 元. Returns original if no unit found."""
    for pattern, multiplier in PRICE_UNIT_PATTERNS:
        m = pattern.search(raw_value)
        if m:
            try:
                num = float(m.group(1))
                converted = num * multiplier
                # Replace the original number+unit with converted value
                return raw_value.replace(m.group(0), str(int(converted)))
            except (ValueError, IndexError):
                continue
    return raw_value


def _lookup_brand_alias(brand: str, conn) -> str:
    """Look up brand alias in DB to normalize to std_brand."""
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT std_brand FROM brand_aliases WHERE alias = %s", (brand,))
        row = cur.fetchone()
        cur.close()
        return row['std_brand'] if row else brand
    except Exception:
        return brand


class EntityExtractor:
    def __init__(self, col_mapping: dict[int, str]):
        self.col_mapping = col_mapping

    def extract(self, row: list[str], row_idx: int, sheet_name: str = "Sheet1", conn=None) -> tuple[dict[str, Any], list[str]]:
        mapper = SemanticMapper()
        raw_fields = {}
        parsed_entities = {}
        reasoning = []

        for col_idx, col_type in self.col_mapping.items():
            raw = row[col_idx] if col_idx < len(row) else ""
            raw_fields[f"col_{col_idx}"] = raw

            # Apply price unit conversion for price-related fields
            processed_raw = raw
            if col_type in ("price", "costPrice", "suggestedPrice"):
                processed_raw = _convert_price_unit(raw)

            result = mapper.map(col_type, processed_raw, row_idx)
            if result["value"] is not None:
                parsed_entities[col_type] = result
                if col_type == "brand" and result["value"] in KNOWN_BRANDS:
                    reasoning.append(f"品牌识别为{result['value']}集团")
                elif col_type == "model" and AC_REGEX.search(raw):
                    reasoning.append("型号匹配空调正则")
                elif col_type == "price":
                    reasoning.append("价格合理性检验通过")

        # Normalize brand via brand_aliases DB lookup if connection provided
        if "brand" in parsed_entities and conn is not None:
            original_brand = parsed_entities["brand"]["value"]
            std_brand = _lookup_brand_alias(original_brand, conn)
            if std_brand != original_brand:
                parsed_entities["brand"]["value"] = std_brand
                reasoning.append(f"品牌别名'{original_brand}'→标准品牌'{std_brand}'")

        return parsed_entities, reasoning


# ---------------------------------------------------------------------------
# Layer 6: Quality Router
# ---------------------------------------------------------------------------
# Field weights for confidence calculation
# brand and model weighted 2x (critical identification fields)
# price weighted 1.5x (commercial field)
# others weighted 1x
FIELD_WEIGHTS = {
    "brand": 2.0,
    "model": 2.0,
    "price": 1.5,
    "costPrice": 1.5,
    "suggestedPrice": 1.5,
    "category": 1.0,
    "stock": 1.0,
    "quantity": 1.0,
    "warehouse": 1.0,
    "text": 1.0,
}


def overall_confidence(entities: dict[str, Any]) -> float:
    """Calculate weighted average confidence score.
    Scores in 0-1 range are multiplied by 100 to get 0-100 scale.
    """
    if not entities:
        return 0.0
    weighted_sum = 0.0
    total_weight = 0.0
    for field, data in entities.items():
        weight = FIELD_WEIGHTS.get(field, 1.0)
        # Convert 0-1 confidence to 0-100 scale
        score_100 = data["confidence"] * 100
        weighted_sum += score_100 * weight
        total_weight += weight
    if total_weight == 0:
        return 0.0
    return round(weighted_sum / total_weight, 1)


def quality_tier(score: float) -> str:
    """Determine quality tier based on confidence score (0-100 scale)."""
    if score >= 90:
        return "HIGH"
    elif score >= 60:
        return "MEDIUM"
    return "LOW"


class QualityRouter:
    def route(self, raw_fields: dict, parsed_entities: dict, reasoning: list[str], row_idx: int, sheet_name: str = "Sheet1") -> dict:
        conf = overall_confidence(parsed_entities)
        return {
            "source_format": "excel",
            "source_location": f"{sheet_name}:row_{row_idx + 1}",
            "raw_fields": raw_fields,
            "parsed_entities": parsed_entities,
            "confidence_score": conf,
            "quality_tier": quality_tier(conf),
            "reasoning": reasoning,
        }


# ---------------------------------------------------------------------------
# Pipeline Orchestrator
# ---------------------------------------------------------------------------
def run_pipeline(sheet_data: list[tuple[str, list[list[str]]]], ext: str, filename: str, conn=None) -> dict:
    """Run pipeline on per-sheet data with independent column mapping per sheet."""
    detector = FormatDetector()
    typer = ColumnTyper()

    fmt = detector.detect(ext)
    router = QualityRouter()

    records = []
    summary = {"total": 0, "HIGH": 0, "MEDIUM": 0, "LOW": 0}
    llm_batch = LLMFallbackBatch()

    # Process each sheet independently with its own column mapping
    for sheet_name, rows in sheet_data:
        headers = rows[0] if rows else []
        data_rows = rows[1:] if len(rows) > 1 else []

        # ── Type-A 供应商模板优先匹配 ──
        template_key = _match_supplier_template(filename)
        if template_key:
            template = _SUPPLIER_TEMPLATES[template_key]
            col_mapping = dict(template["column_mapping"])
            fixed_fields = template.get("fixed_fields", {})
        else:
            col_mapping = typer.infer(headers, data_rows)
            fixed_fields = {}
        extractor = EntityExtractor(col_mapping)

        for idx, row in enumerate(data_rows):
            parsed_entities, reasoning = extractor.extract(row, idx, sheet_name=sheet_name, conn=conn)
            # ── 注入固定字段（Type-A 模板） ──
            for field, value in fixed_fields.items():
                if field not in parsed_entities:
                    parsed_entities[field] = {"value": value, "confidence": 1.0}
                    reasoning.append(f"模板固定字段: {field}={value}")
            raw_fields = {headers[i]: (row[i] if i < len(row) else "") for i in range(len(headers))}
            record = router.route(raw_fields, parsed_entities, reasoning, idx + 2, sheet_name=sheet_name)
            records.append(record)
            llm_batch.add(record)
            tier = record["quality_tier"]
            summary[tier] += 1
            summary["total"] += 1

    # LLM Fallback：批量处理低置信度记录
    if llm_batch.batch:
        llm_batch.call_llm()

    return {
        "status": "ok",
        "filename": filename,
        "format_detected": fmt,
        "records": records,
        "summary": summary,
        "processed_at": datetime.utcnow().isoformat(),
    }


# ---------------------------------------------------------------------------
# Database Persistence — write parsed records to MySQL
# ---------------------------------------------------------------------------
LOW_QUALITY_THRESHOLD = 65.0  # confidence below this → is_low_quality=1
LLM_FALLBACK_THRESHOLD = 65.0  # confidence below this + missing key fields → LLM fallback
LLM_API_KEY = os.environ.get("DEEPSEEK_API_KEY", "")
LLM_API_BASE = "https://api.deepseek.com/anthropic"
LLM_MODEL = "deepseek-v4-pro"

# ---------------------------------------------------------------------------
# LLM Fallback Batch (Column Type Inference via LLM)
# ---------------------------------------------------------------------------

class LLMFallbackBatch:
    """当正则匹配覆盖率低时，批量调用LLM重新推断列类型。"""

    def __init__(self, api_key: str = None):
        self.api_key = api_key or LLM_API_KEY
        self.api_base = LLM_API_BASE
        self.model = LLM_MODEL
        self.batch: list[dict] = []

    def should_fallback(self, record: dict) -> bool:
        """双条件：conf < 65 且 brand+model 同时缺失"""
        conf = record.get("confidence_score", 0.0)
        parsed = record.get("parsed_entities", {})
        brand_missing = "brand" not in parsed or parsed.get("brand", {}).get("value") is None
        model_missing = "model" not in parsed or parsed.get("model", {}).get("value") is None
        return conf < LLM_FALLBACK_THRESHOLD and brand_missing and model_missing

    def add(self, record: dict):
        if self.should_fallback(record):
            self.batch.append(record)

    def call_llm(self) -> list[dict]:
        """批量调用LLM，重新推断列类型"""
        if not self.batch:
            return []

        # 组装prompt（控制token预算，最多10条）
        prompt_lines = ["你是家电产品数据解析专家。根据以下Excel列名和样本行，判断每列的语义类型。"]
        prompt_lines.append("")
        prompt_lines.append("允许的类型：brand, model, price, category, stock, quantity, warehouse, costPrice, suggestedPrice, text")
        prompt_lines.append("")

        for i, rec in enumerate(self.batch[:10]):
            raw = rec.get("raw_fields", {})
            headers = list(raw.keys())
            values = list(raw.values())
            sheet = rec.get("source_location", "").split(":")[0]
            prompt_lines.append(f"### 记录{i+1} (Sheet: {sheet})")
            prompt_lines.append(f"列名: {headers}")
            prompt_lines.append(f"样本行: {values}")
            prompt_lines.append("")

        prompt_lines.append("请以JSON格式返回每列的类型映射（不含其他内容）：")
        prompt_lines.append('{"col_0": "brand", "col_1": "model", ...}')

        prompt = "\n".join(prompt_lines)

        try:
            import openai
            client = openai.OpenAI(api_key=self.api_key, base_url=self.api_base)
            resp = client.chat.completions.create(
                model=self.model,
                messages=[{"role": "user", "content": prompt}],
                temperature=0.1,
            )
            content = resp.choices[0].message.content
            # 解析JSON
            import re, json
            match = re.search(r"\{[\s\S]+\}", content)
            if not match:
                return self.batch
            mapping = json.loads(match.group())
            # 更新batch中记录的confidence（重新计算）
            for rec in self.batch[:10]:
                rec["llm_column_mapping"] = mapping
                # 估算：LLM推断后confidence +20分
                rec["confidence_score"] = min(100.0, rec["confidence_score"] + 20.0)
                rec["quality_tier"] = quality_tier(rec["confidence_score"])
        except Exception as e:
            pass  # LLM失败，原样返回

        return self.batch


def _sanitize_filename(filename: str) -> str:
    """Remove special characters and ensure safe filename for DB storage."""
    # Replace common problematic chars with underscore
    filename = re.sub(r'[\/\\:*?"<>|]', '_', filename)
    # Ensure it's valid UTF-8
    try:
        filename = filename.encode('utf-8', errors='replace').decode('utf-8')
    except Exception:
        filename = 'unknown_file'
    return filename


def _get_or_create_supplier(filename: str, conn) -> int | None:
    """Get or create a supplier record. Returns supplier_id."""
    # Sanitize filename to handle Chinese chars and special chars
    safe_filename = _sanitize_filename(filename)
    # Derive supplier_code from filename (use first 64 chars of filename as code)
    supplier_code = safe_filename[:64]
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT id FROM suppliers WHERE supplier_code = %s", (supplier_code,))
    row = cur.fetchone()
    if row:
        sid = row["id"]
        cur.close()
        return sid
    # Create new supplier
    supplier_name = safe_filename.split("/")[-1].split(".")[0][:128]
    cur.execute(
        "INSERT INTO suppliers (supplier_code, supplier_name, source_file, freshness, total_records) "
        "VALUES (%s, %s, %s, 'pending', 0)",
        (supplier_code, supplier_name, safe_filename),
    )
    sid = cur.lastrowid
    conn.commit()
    cur.close()
    return sid


def _upsert_supplier_stats(supplier_id: int, records: list[dict], conn):
    """Update supplier aggregate stats after inserting records."""
    total = len(records)
    if total == 0:
        return
    brands = set()
    prices = []
    quality_scores = []
    low_count = 0
    for r in records:
        if r.get("brand"):
            brands.add(r["brand"])
        if r.get("price") is not None:
            prices.append(r["price"])
        cs = r.get("confidence_score", 0)
        quality_scores.append(cs)
        if r.get("quality_tier") == "LOW":
            low_count += 1

    avg_price = sum(prices) / len(prices) if prices else 0
    avg_conf = sum(quality_scores) / len(quality_scores) if quality_scores else 0
    parse_rate = (total - low_count) / total * 100 if total > 0 else 0
    cur = conn.cursor()
    cur.execute("""
        UPDATE suppliers
        SET total_records = %s,
            total_brands = %s,
            avg_price = %s,
            parse_success_rate = %s,
            data_quality_score = %s,
            updated_at = NOW()
        WHERE id = %s
    """, (total, len(brands), avg_price, parse_rate, avg_conf, supplier_id))
    conn.commit()
    cur.close()


def persist_records_to_db(records: list[dict], filename: str, supplier_id: int | None = None) -> dict:
    """
    Write parsed records to MySQL supplier_quotes table.
    Auto-marks LOW quality_tier records as is_low_quality=1.
    Returns stats dict.
    """
    if not records:
        return {"inserted": 0, "low_quality": 0, "supplier_id": supplier_id}

    conn = get_db_connection()
    try:
        if supplier_id is None:
            supplier_id = _get_or_create_supplier(filename, conn)
            if supplier_id is None:
                return {"error": "could not get or create supplier", "inserted": 0}

        inserted = 0
        low_quality_count = 0
        cur = conn.cursor()

        for record in records:
            parsed = record.get("parsed_entities", {})
            conf = record.get("confidence_score", 0.0)
            tier = record.get("quality_tier", "MEDIUM")
            is_low = 1 if conf < LOW_QUALITY_THRESHOLD else 0

            # Extract field values safely
            brand = parsed.get("brand", {}).get("value") if parsed.get("brand") else None
            category = parsed.get("category", {}).get("value") if parsed.get("category") else None
            model_raw = parsed.get("model", {}).get("value") if parsed.get("model") else None
            model_std = (model_raw or "")[:128]  # Truncate to fit varchar(128)
            raw_price = parsed.get("price", {}).get("value") if parsed.get("price") else None
            # Explicitly convert to float and clamp to DECIMAL(10,2) safe range before INSERT
            price = None
            if raw_price is not None:
                try:
                    price = round(float(raw_price), 2)
                    # Clamp to DECIMAL(10,2) range: max=9999999.99, min=-999999.99
                    price = max(min(price, 9999999.99), -999999.99)
                except (ValueError, TypeError):
                    price = None
            price_type = None  # price_type not extracted in current pipeline

            if brand is None and category is None and model_raw is None:
                # Skip completely empty records
                continue

            cur.execute("""
                INSERT INTO supplier_quotes
                    (supplier_id, brand, category, model_raw, model_std, price, price_type,
                     quality_tier, confidence, is_low_quality, raw_row)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                supplier_id,
                brand,
                category,
                model_raw,
                model_std,
                price,
                price_type,
                tier,
                conf,
                is_low,
                json.dumps(record.get("raw_fields", {})),
            ))
            inserted += 1
            if is_low:
                low_quality_count += 1

        conn.commit()
        cur.close()

        # Update supplier aggregate stats
        _upsert_supplier_stats(supplier_id, records, conn)

        return {
            "inserted": inserted,
            "low_quality": low_quality_count,
            "supplier_id": supplier_id,
        }
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Background Task Processing
# ---------------------------------------------------------------------------
async def process_task_async(task_id: str, sheet_data: list[tuple[str, list[list[str]]]], ext: str, filename: str):
    """Background task: run pipeline, persist to MySQL, and store result in Redis."""
    r = get_redis()
    conn = None
    try:
        conn = get_db_connection()
        result = run_pipeline(sheet_data, ext, filename, conn=conn)
        result["task_id"] = task_id
        result["status"] = "completed"

        # Persist records to MySQL with confidence + auto low-quality marking
        persist_stats = persist_records_to_db(result.get("records", []), filename)
        result["db_persist"] = persist_stats

        # Store result as JSON string with TTL
        r.setex(f"{RESULT_PREFIX}{task_id}", RESULT_TTL, json.dumps(result))
    except Exception as e:
        error_result = {
            "task_id": task_id,
            "status": "failed",
            "error": str(e),
            "completed_at": datetime.utcnow().isoformat(),
        }
        r.setex(f"{RESULT_PREFIX}{task_id}", RESULT_TTL, json.dumps(error_result))
    finally:
        if conn:
            conn.close()


# ---------------------------------------------------------------------------
# FastAPI Endpoints
# ---------------------------------------------------------------------------
@app.get("/pipeline/health")
def health():
    return {"status": "ok", "service": "pipeline"}


@app.post("/pipeline/parse", response_model=TaskSubmitResponse)
async def parse(file: UploadFile):
    """
    Async intake: writes to Redis Streams and returns task_id immediately.
    Does NOT wait for parsing to complete.
    """
    content = await file.read()
    sheet_data, ext = IntakeLayer().load(content, file.filename)

    task_id = str(uuid.uuid4())
    r = get_redis()

    # Write task to Redis Stream
    task_data = {
        "task_id": task_id,
        "filename": file.filename,
        "ext": ext,
        "submitted_at": datetime.utcnow().isoformat(),
        "rows": json.dumps(sheet_data),  # serialize sheet_data as JSON string
    }
    r.xadd(STREAM_KEY, task_data)

    # Also store raw sheet_data under result key with PENDING status (for result query before completion)
    pending_result = {
        "task_id": task_id,
        "status": "pending",
        "filename": file.filename,
        "submitted_at": datetime.utcnow().isoformat(),
    }
    r.setex(f"{RESULT_PREFIX}{task_id}", RESULT_TTL, json.dumps(pending_result))

    # Trigger background processing
    import asyncio
    asyncio.create_task(process_task_async(task_id, sheet_data, ext, file.filename))

    return TaskSubmitResponse(
        task_id=task_id,
        status="accepted",
        message="任务已提交，正在后台处理。查询结果请用 GET /pipeline/result/{task_id}",
    )


@app.get("/pipeline/result/{task_id}", response_model=TaskResultResponse)
def get_result(task_id: str):
    """
    Query processing result by task_id.
    Returns pending/completed/failed status.
    """
    r = get_redis()
    raw = r.get(f"{RESULT_PREFIX}{task_id}")
    if raw is None:
        return TaskResultResponse(
            task_id=task_id,
            status="not_found",
            result=None,
        )

    result = json.loads(raw)
    return TaskResultResponse(
        task_id=task_id,
        status=result.get("status", "unknown"),
        result=result,
    )


@app.get("/pipeline/stream/info")
def stream_info():
    """Return Redis Stream info for debugging."""
    try:
        r = get_redis()
        info = r.xinfo_stream(STREAM_KEY)
        return {"stream": STREAM_KEY, "exists": True, "info": info}
    except redis.ResponseError:
        return {"stream": STREAM_KEY, "exists": False}

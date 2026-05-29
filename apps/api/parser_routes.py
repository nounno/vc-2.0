"""
Parser Management API Routes
Covers: upload, job management, row preview, commit.
"""
import json
import math
import os
import uuid
from datetime import datetime
from typing import Optional, List, Any

from fastapi import APIRouter, UploadFile, File, Form, HTTPException, Query, Depends, BackgroundTasks
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from app.auth import get_db, require_auth

router = APIRouter(prefix="/api/v1/parser", tags=["parser"])

# -----------------------------------------------------------------------------
# Pydantic models
# ---------------------------------------------------------------------------
class SupplierSelect(BaseModel):
    supplier_id: Optional[int] = None
    supplier_name: Optional[str] = None

class ConfirmStageRequest(BaseModel):
    stage: str

class RowModifyData(BaseModel):
    brand: Optional[str] = None
    category: Optional[str] = None
    model: Optional[str] = None
    model_std: Optional[str] = None
    price: Optional[float] = None
    description: Optional[str] = None
    notes: Optional[str] = None

class RowUpdateRequest(BaseModel):
    action: str  # "modify" | "reject" | "split"
    data: Optional[RowModifyData] = None
    splits: Optional[List[dict]] = None

class CommitRequest(BaseModel):
    action: str  # "commit_all" | "commit_valid"

# ---------------------------------------------------------------------------
# Helper: generate job code
# ---------------------------------------------------------------------------
def _generate_job_code() -> str:
    today = datetime.now().strftime("%Y%m%d")
    import random
    seq = random.randint(1, 999)
    return f"P-{today}-{seq:03d}"

# ---------------------------------------------------------------------------
# Helper: get next stage
# ---------------------------------------------------------------------------
STAGE_ORDER = ["uploaded", "split", "cleaned", "mapped", "standardized", "committed"]

def _next_stage(current: str) -> Optional[str]:
    try:
        idx = STAGE_ORDER.index(current)
        if idx + 1 < len(STAGE_ORDER):
            return STAGE_ORDER[idx + 1]
    except ValueError:
        pass
    return None

# ---------------------------------------------------------------------------
# 1. Upload + start parsing
# ---------------------------------------------------------------------------
@router.post("/upload")
async def parser_upload(
    current_user: dict = Depends(require_auth()),
    file: UploadFile = File(...),
    supplier_id: Optional[int] = Form(None),
    supplier_name: Optional[str] = Form(None),
):
    """
    Upload Excel file, optionally create/select supplier,
    auto-process through split→clean→map→standardize,
    store results in parse_jobs + parse_row_staging.
    """
    filename = file.filename or ""
    if not filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "只支持 .xlsx / .xls 文件")

    file_bytes = await file.read()

    # Resolve supplier
    with get_db() as db:
        with db.cursor() as cursor:
            if supplier_id:
                cursor.execute("SELECT id, supplier_name FROM supplier_files WHERE id=%s", (supplier_id,))
                row = cursor.fetchone()
                if not row:
                    raise HTTPException(404, "供应商不存在")
                supplier_name_val = row["supplier_name"]
            elif supplier_name:
                # Check if exists
                cursor.execute("SELECT id FROM supplier_files WHERE supplier_name=%s LIMIT 1", (supplier_name,))
                existing = cursor.fetchone()
                if existing:
                    supplier_id = existing["id"]
                    supplier_name_val = supplier_name
                else:
                    # Create new supplier
                    supplier_code = f"S-{datetime.now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:6]}"
                    cursor.execute(
                        "INSERT INTO supplier_files (supplier_code, supplier_name) VALUES (%s, %s)",
                        (supplier_code, supplier_name),
                    )
                    db.commit()
                    supplier_id = cursor.lastrowid
                    supplier_name_val = supplier_name
            else:
                raise HTTPException(400, "必须提供 supplier_id 或 supplier_name")

    # Generate job code
    job_code = _generate_job_code()

    # Save file to disk
    upload_dir = "/tmp/parser_uploads"
    os.makedirs(upload_dir, exist_ok=True)
    safe_filename = filename.replace("/", "_").replace("\\", "_")
    file_path = os.path.join(upload_dir, f"{job_code}_{safe_filename}")
    with open(file_path, "wb") as f:
        f.write(file_bytes)

    # Run parsing pipeline
    # supplier_id is guaranteed non-None here (validated above)
    from parser_core import process_excel
    try:
        result = process_excel(file_bytes, job_code, int(supplier_id))
    except Exception as e:
        raise HTTPException(500, f"解析失败: {str(e)}")

    # Persist to DB
    import json
    with get_db() as db:
        with db.cursor() as cursor:
            cursor.execute("""
                INSERT INTO parse_jobs
                (job_code, supplier_id, original_filename, file_path, status,
                 sheet_count, total_rows, valid_rows, flagged_rows,
                 split_summary, cleaning_summary, mapping_summary, standardization_summary)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                job_code,
                supplier_id,
                file.filename,
                file_path,
                "uploaded",
                result["sheet_count"],
                result["total_rows"],
                result["valid_rows"],
                result["flagged_rows"],
                json.dumps(result["sheets"]),
                json.dumps({"method": "auto"}),
                json.dumps({"method": "auto"}),
                json.dumps({"method": "auto"}),
            ))
            db.commit()
            job_db_id = cursor.lastrowid

            # Insert row staging
            for sheet in result["_raw_sheets"]:
                sheet_name = sheet["name"]
                for r in sheet["rows"]:
                    cursor.execute("""
                        INSERT INTO parse_row_staging
                        (job_id, sheet_name, row_index, row_status, confidence,
                         brand, category, model, model_std, price, description, notes,
                         raw_data, source_columns, confidence_details, is_multi_product)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (
                        job_db_id,
                        sheet_name,
                        r.get("row_num", 0),
                        r.get("status", "valid"),
                        r.get("confidence", 0),
                        r.get("brand"),
                        r.get("category"),
                        r.get("model"),
                        r.get("model_std"),
                        r.get("price"),
                        r.get("description"),
                        r.get("notes"),
                        json.dumps(r.get("raw_data")),
                        json.dumps(r.get("source_columns", {})),
                        json.dumps(r.get("confidence_details", {})),
                        1 if r.get("status") == "pending_split" else 0,
                    ))
            db.commit()

    return {
        "job_id": job_db_id,
        "job_code": job_code,
        "supplier_id": supplier_id,
        "supplier_name": supplier_name_val,
        "status": "uploaded",
        "sheet_count": result["sheet_count"],
        "sheets": result["sheets"],
        "message": "上传成功，请在切分阶段确认后继续",
    }


# ---------------------------------------------------------------------------
# 2. Get job list
# ---------------------------------------------------------------------------
@router.get("/jobs")
def parser_job_list(
    current_user: dict = Depends(require_auth()),
    supplier_id: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
):
    with get_db() as db:
        with db.cursor() as cursor:
            where = ["1=1"]
            params = []
            if supplier_id:
                where.append("pj.supplier_id = %s")
                params.append(supplier_id)
            if status:
                where.append("pj.status = %s")
                params.append(status)

            count_sql = f"SELECT COUNT(*) as total FROM parse_jobs pj WHERE {' AND '.join(where)}"
            cursor.execute(count_sql, params)
            total = cursor.fetchone()["total"]

            offset = (page - 1) * limit
            cursor.execute(f"""
                SELECT pj.id, pj.job_code, pj.supplier_id, pj.original_filename,
                       pj.status, pj.sheet_count, pj.total_rows, pj.valid_rows,
                       pj.flagged_rows, pj.committed_rows, pj.rejected_rows,
                       pj.created_at, pj.updated_at,
                       sf.supplier_name
                FROM parse_jobs pj
                LEFT JOIN supplier_files sf ON pj.supplier_id = sf.id
                WHERE {' AND '.join(where)}
                ORDER BY pj.created_at DESC
                LIMIT %s OFFSET %s
            """, params + [limit, offset])
            jobs = cursor.fetchall()

            # Serialize dates
            for j in jobs:
                if j.get("created_at"):
                    j["created_at"] = j["created_at"].isoformat()
                if j.get("updated_at"):
                    j["updated_at"] = j["updated_at"].isoformat()

            return {"jobs": jobs, "total": total, "page": page, "limit": limit}


# ---------------------------------------------------------------------------
# 3. Get job detail
# ---------------------------------------------------------------------------
@router.get("/jobs/{job_id}")
def parser_job_detail(job_id: int, current_user: dict = Depends(require_auth())):
    with get_db() as db:
        with db.cursor() as cursor:
            cursor.execute("""
                SELECT pj.*, sf.supplier_name
                FROM parse_jobs pj
                LEFT JOIN supplier_files sf ON pj.supplier_id = sf.id
                WHERE pj.id=%s
            """, (job_id,))
            job = cursor.fetchone()
            if not job:
                raise HTTPException(404, "任务不存在")

            if job.get("created_at"):
                job["created_at"] = job["created_at"].isoformat()
            if job.get("updated_at"):
                job["updated_at"] = job["updated_at"].isoformat()

            return job


# ---------------------------------------------------------------------------
# 4. Confirm next stage
# ---------------------------------------------------------------------------
@router.post("/jobs/{job_id}/confirm-stage")
def parser_confirm_stage(job_id: int, body: ConfirmStageRequest, current_user: dict = Depends(require_auth())):
    """
    Move job from current stage to next.
    Stages: uploaded → split → cleaned → mapped → standardized → committed
    """
    with get_db() as db:
        with db.cursor() as cursor:
            cursor.execute("SELECT id, status FROM parse_jobs WHERE id=%s", (job_id,))
            job = cursor.fetchone()
            if not job:
                raise HTTPException(404, "任务不存在")

            current = job["status"]
            next_s = _next_stage(current)

            if body.stage != next_s:
                raise HTTPException(
                    409,
                    f"当前阶段是 {current}，下一阶段应该是 {next_s}，不是 {body.stage}"
                )

            cursor.execute(
                "UPDATE parse_jobs SET status=%s, updated_at=NOW() WHERE id=%s",
                (body.stage, job_id),
            )
            db.commit()

            return {"job_id": job_id, "status": body.stage, "next_stage": _next_stage(body.stage)}


# ---------------------------------------------------------------------------
# 4b. Get sheets (切分结果)
# ---------------------------------------------------------------------------
@router.get("/jobs/{job_id}/sheets")
def parser_sheets(job_id: int, current_user: dict = Depends(require_auth())):
    """
    返回各 Sheet 的切分结果：名称、行数、状态。
    用户在阶段一确认后调用 confirm-split 进入阶段二。
    """
    with get_db() as db:
        with db.cursor() as cursor:
            cursor.execute(
                "SELECT id, status, sheet_count, total_rows, valid_rows, flagged_rows, split_summary "
                "FROM parse_jobs WHERE id=%s",
                (job_id,)
            )
            job = cursor.fetchone()
            if not job:
                raise HTTPException(404, "任务不存在")

            import json
            sheets_raw = job.get("split_summary") or "[]"
            if isinstance(sheets_raw, str):
                sheets = json.loads(sheets_raw)
            else:
                sheets = sheets_raw

            return {
                "job_id": job_id,
                "status": job["status"],
                "sheet_count": job["sheet_count"],
                "total_rows": job["total_rows"],
                "sheets": sheets,
            }


# ---------------------------------------------------------------------------
# 4c. Rollback job to previous stage
# ---------------------------------------------------------------------------
@router.post("/jobs/{job_id}/rollback")
def parser_rollback(job_id: int, body: ConfirmStageRequest, current_user: dict = Depends(require_auth())):
    """
    将任务回退到指定阶段（只能回退到上一阶段）。
    """
    with get_db() as db:
        with db.cursor() as cursor:
            cursor.execute(
                "SELECT id, status FROM parse_jobs WHERE id=%s",
                (job_id,)
            )
            job = cursor.fetchone()
            if not job:
                raise HTTPException(404, "任务不存在")

            current = job["status"]
            # Rollback: go back one stage in STAGE_ORDER
            try:
                idx = STAGE_ORDER.index(current)
                if idx == 0:
                    raise HTTPException(409, "已在最开始的阶段，无法回退")
                target_stage = STAGE_ORDER[idx - 1]
            except ValueError:
                raise HTTPException(500, f"未知状态: {current}")

            cursor.execute(
                "UPDATE parse_jobs SET status=%s, updated_at=NOW() WHERE id=%s",
                (target_stage, job_id),
            )
            db.commit()

            return {
                "job_id": job_id,
                "previous_status": current,
                "current_status": target_stage,
                "message": f"已回退至 {target_stage}",
            }


# ---------------------------------------------------------------------------
# 5. Get preview rows
# ---------------------------------------------------------------------------
@router.get("/jobs/{job_id}/rows")
def parser_rows(
    job_id: int,
    current_user: dict = Depends(require_auth()),
    status: Optional[str] = Query(None),
    sheet: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=200),
):
    with get_db() as db:
        with db.cursor() as cursor:
            # Verify job exists
            cursor.execute("SELECT id FROM parse_jobs WHERE id=%s", (job_id,))
            if not cursor.fetchone():
                raise HTTPException(404, "任务不存在")

            where = ["job_id=%s"]
            params = [job_id]
            if status and status != "all":
                where.append("row_status=%s")
                params.append(status)
            if sheet:
                where.append("sheet_name=%s")
                params.append(sheet)

            count_sql = f"SELECT COUNT(*) as total FROM parse_row_staging WHERE {' AND '.join(where)}"
            cursor.execute(count_sql, params)
            total = cursor.fetchone()["total"]

            offset = (page - 1) * limit
            cursor.execute(f"""
                SELECT id, sheet_name, row_index, row_status, confidence,
                       brand, category, model, model_std, price,
                       description, notes, raw_data, source_columns,
                       confidence_details, is_multi_product, split_into,
                       manually_corrected, reviewer_action, created_at, updated_at
                FROM parse_row_staging
                WHERE {' AND '.join(where)}
                ORDER BY
                    CASE WHEN row_status='flagged' THEN 0
                         WHEN row_status='pending_split' THEN 1
                         WHEN row_status='valid' THEN 2
                         ELSE 3 END,
                    confidence ASC,
                    row_index ASC
                LIMIT %s OFFSET %s
            """, params + [limit, offset])
            rows = cursor.fetchall()

            for r in rows:
                if r.get("created_at"):
                    r["created_at"] = r["created_at"].isoformat()
                if r.get("updated_at"):
                    r["updated_at"] = r["updated_at"].isoformat()

            return {"rows": rows, "total": total, "page": page, "limit": limit}


# ---------------------------------------------------------------------------
# 6. Update row (modify/reject/split)
# ---------------------------------------------------------------------------
@router.patch("/rows/{row_id}")
def parser_update_row(row_id: int, body: RowUpdateRequest, current_user: dict = Depends(require_auth())):
    with get_db() as db:
        with db.cursor() as cursor:
            cursor.execute("SELECT id, job_id, row_status FROM parse_row_staging WHERE id=%s", (row_id,))
            row = cursor.fetchone()
            if not row:
                raise HTTPException(404, "行记录不存在")

            job_id = row["job_id"]
            action = body.action

            if action == "reject":
                cursor.execute(
                    "UPDATE parse_row_staging SET row_status='rejected', reviewer_action='reject', "
                    "updated_at=NOW() WHERE id=%s",
                    (row_id,),
                )
                new_status = "rejected"

            elif action == "modify":
                data = body.data
                if not data:
                    raise HTTPException(400, "修改数据不能为空")
                updates = []
                params = []
                for field in ["brand", "category", "model", "model_std",
                               "price", "description", "notes"]:
                    val = getattr(data, field, None)
                    if val is not None:
                        updates.append(f"{field}=%s")
                        params.append(val)
                if not updates:
                    raise HTTPException(400, "没有有效字段要修改")
                updates.append("manually_corrected=1")
                updates.append("reviewer_action='modify'")
                params.append(row_id)
                cursor.execute(
                    f"UPDATE parse_row_staging SET {', '.join(updates)}, updated_at=NOW() WHERE id=%s",
                    params,
                )
                new_status = "valid"  # after modification, treat as valid

            elif action == "split":
                if not body.splits:
                    raise HTTPException(400, "拆分数据不能为空")
                # Mark current row as pending_split and store splits
                import json
                cursor.execute(
                    "UPDATE parse_row_staging SET is_multi_product=1, row_status='pending_split', "
                    "split_into=%s, reviewer_action='split', updated_at=NOW() WHERE id=%s",
                    (json.dumps(body.splits), row_id),
                )
                new_status = "pending_split"

            else:
                raise HTTPException(400, f"未知操作: {action}")

            db.commit()

            # Recalculate confidence after modification
            if action == "modify":
                cursor.execute(
                    "UPDATE parse_row_staging SET confidence=90.0, updated_at=NOW() WHERE id=%s",
                    (row_id,),
                )
                db.commit()

            return {"row_id": row_id, "new_status": new_status, "action": action}


# ---------------------------------------------------------------------------
# 7. Commit to supplier_quotes
# ---------------------------------------------------------------------------
@router.post("/jobs/{job_id}/commit")
def parser_commit(job_id: int, body: CommitRequest, current_user: dict = Depends(require_auth())):
    """
    Write staged rows to supplier_quotes.
    action=commit_all: import all non-rejected rows
    action=commit_valid: import only valid rows (skip flagged/pending_split)
    """
    with get_db() as db:
        with db.cursor() as cursor:
            cursor.execute("SELECT id, status, supplier_id FROM parse_jobs WHERE id=%s", (job_id,))
            job = cursor.fetchone()
            if not job:
                raise HTTPException(404, "任务不存在")

            supplier_id = job["supplier_id"]

            # Determine which rows to import
            if body.action == "commit_valid":
                where_clause = "job_id=%s AND row_status IN ('valid')"
            else:  # commit_all
                where_clause = "job_id=%s AND row_status NOT IN ('rejected')"

            cursor.execute(f"SELECT COUNT(*) as cnt FROM parse_row_staging WHERE {where_clause}", (job_id,))
            to_import = cursor.fetchone()["cnt"]

            # Fetch rows to import
            cursor.execute(f"""
                SELECT brand, category, model, model_std, price, description, notes,
                       row_status, confidence
                FROM parse_row_staging
                WHERE {where_clause}
            """, (job_id,))
            rows_to_import = cursor.fetchall()

            # Insert into supplier_quotes
            imported = 0
            for r in rows_to_import:
                try:
                    # Determine quality_tier and is_low_quality from row_status
                    row_status = r.get("row_status", "valid")
                    row_confidence = float(r.get("confidence") or 0)
                    if row_status == "flagged":
                        quality_tier = "LOW"
                        is_low_quality = 1
                    elif row_status == "rejected":
                        quality_tier = "LOW"
                        is_low_quality = 1
                    else:
                        quality_tier = "MEDIUM"
                        is_low_quality = 0

                    cursor.execute("""
                        INSERT INTO supplier_quotes
                        (supplier_id, brand, category, model_raw, model_std, price,
                         quality_tier, is_low_quality, confidence, raw_row, created_at)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
                    """, (
                        supplier_id,
                        r.get("brand"),
                        r.get("category"),
                        r.get("model"),
                        r.get("model_std"),
                        r.get("price"),
                        quality_tier,
                        is_low_quality,
                        row_confidence,
                        '{"source": "parser_staging"}',
                    ))
                    imported += 1
                except Exception:
                    pass  # skip duplicates or bad data

            # Update job stats
            cursor.execute("""
                UPDATE parse_jobs
                SET status='committed',
                    committed_rows=%s,
                    rejected_rows=(SELECT COUNT(*) FROM parse_row_staging WHERE job_id=%s AND row_status='rejected'),
                    updated_at=NOW()
                WHERE id=%s
            """, (imported, job_id, job_id))
            db.commit()

            return {
                "job_id": job_id,
                "committed": imported,
                "message": f"成功入库 {imported} 条记录"
            }


# --------------------------------------------------------------------------
# 8. Report endpoints
# --------------------------------------------------------------------------


def _job_exists(job_id: int):
    """Check if job exists, raise 404 if not."""
    with get_db() as db:
        with db.cursor() as cursor:
            cursor.execute("SELECT id, status, job_code FROM parse_jobs WHERE id=%s", (job_id,))
            job = cursor.fetchone()
            if not job:
                raise HTTPException(404, "任务不存在")
            return job


@router.get("/jobs/{job_id}/report")
def parser_report(job_id: int, current_user: dict = Depends(require_auth())):
    """
    总览报告：聚合 split/cleaning/mapping/standardization 四个阶段的统计。
    """
    job = _job_exists(job_id)

    with get_db() as db:
        with db.cursor() as cursor:
            # Get row counts by status
            cursor.execute("""
                SELECT row_status, COUNT(*) as cnt
                FROM parse_row_staging
                WHERE job_id=%s
                GROUP BY row_status
            """, (job_id,))
            status_counts = {r["row_status"]: r["cnt"] for r in cursor.fetchall()}

            total_rows = sum(status_counts.values())
            valid = status_counts.get("valid", 0)
            flagged = status_counts.get("flagged", 0)
            rejected = status_counts.get("rejected", 0)
            pending_split = status_counts.get("pending_split", 0)

            # Get sheets info
            cursor.execute("""
                SELECT sheet_name, COUNT(*) as cnt
                FROM parse_row_staging
                WHERE job_id=%s
                GROUP BY sheet_name
            """, (job_id,))
            sheets = [{"name": r["sheet_name"], "rows": r["cnt"]} for r in cursor.fetchall()]

            # Get avg confidence and model_std fill count
            cursor.execute("""
                SELECT
                    AVG(confidence) as avg_conf,
                    SUM(CASE WHEN model_std IS NOT NULL AND model_std != '' THEN 1 ELSE 0 END) as model_std_filled,
                    COUNT(*) as total
                FROM parse_row_staging
                WHERE job_id=%s
            """, (job_id,))
            stats = cursor.fetchone()

            avg_conf = float(stats["avg_conf"]) if stats["avg_conf"] else 0.0
            model_std_filled = stats["model_std_filled"] or 0

            # Count distinct mapped fields (read source_columns JSON and count unique keys)
            cursor.execute("""
                SELECT source_columns
                FROM parse_row_staging
                WHERE job_id=%s AND source_columns IS NOT NULL AND source_columns != ''
            """, (job_id,))
            _all_sc_rows = cursor.fetchall()
            _seen_keys = set()
            for _r in _all_sc_rows:
                try:
                    _sc = _r["source_columns"]
                    if isinstance(_sc, str):
                        _sc = json.loads(_sc)
                    _seen_keys.update(_sc.keys())
                except Exception:
                    pass
            fields_mapped = len(_seen_keys)

    import json as _json

    # Get split_summary for sheets
    split_summary = []
    try:
        with get_db() as db:
            with db.cursor() as cursor:
                cursor.execute("SELECT split_summary FROM parse_jobs WHERE id=%s", (job_id,))
                raw = cursor.fetchone()["split_summary"]
                if raw:
                    if isinstance(raw, str):
                        split_summary = _json.loads(raw)
                    else:
                        split_summary = raw
    except Exception:
        split_summary = []

    return {
        "job_id": job_id,
        "job_code": job["job_code"],
        "status": job["status"],
        "stages": {
            "split": {
                "status": "done",
                "total_rows": total_rows,
                "sheets": sheets if sheets else split_summary,
            },
            "cleaning": {
                "status": "done",
                "valid": valid,
                "flagged": flagged,
                "rejected": rejected,
                "pending_split": pending_split,
            },
            "mapping": {
                "status": "done",
                "fields_mapped": fields_mapped,
            },
            "standardization": {
                "status": "done",
                "avg_confidence": round(avg_conf, 1),
                "model_std_filled": model_std_filled,
            },
        },
    }


@router.get("/jobs/{job_id}/report/cleaning")
def parser_report_cleaning(job_id: int, current_user: dict = Depends(require_auth())):
    """
    清洗报告：按 row_status 分组统计行数和百分比，以及 confidence 分布。
    """
    _job_exists(job_id)

    with get_db() as db:
        with db.cursor() as cursor:
            # Status counts
            cursor.execute("""
                SELECT row_status, COUNT(*) as cnt
                FROM parse_row_staging
                WHERE job_id=%s
                GROUP BY row_status
            """, (job_id,))
            status_rows = cursor.fetchall()
            total = sum(r["cnt"] for r in status_rows)

            status_distribution = []
            for r in status_rows:
                pct = round(r["cnt"] / total * 100, 2) if total > 0 else 0.0
                status_distribution.append({
                    "status": r["row_status"],
                    "count": r["cnt"],
                    "percentage": pct,
                })

            # Confidence distribution
            cursor.execute("""
                SELECT
                    MIN(confidence) as min_conf,
                    MAX(confidence) as max_conf,
                    AVG(confidence) as avg_conf,
                    COUNT(*) as total
                FROM parse_row_staging
                WHERE job_id=%s
            """, (job_id,))
            conf_stats = cursor.fetchone()

            # Confidence percentiles
            cursor.execute("""
                SELECT confidence
                FROM parse_row_staging
                WHERE job_id=%s
                ORDER BY confidence ASC
            """, (job_id,))
            all_confs = [row["confidence"] for row in cursor.fetchall()]
            n = len(all_confs)

            def percentile(vals, p):
                if not vals:
                    return 0.0
                idx = int((p / 100.0) * (len(vals) - 1))
                return float(vals[idx])

            p25 = percentile(all_confs, 25)
            p50 = percentile(all_confs, 50)
            p75 = percentile(all_confs, 75)

    return {
        "job_id": job_id,
        "total_rows": total,
        "status_distribution": status_distribution,
        "confidence_distribution": {
            "min": float(conf_stats["min_conf"]) if conf_stats["min_conf"] else 0.0,
            "max": float(conf_stats["max_conf"]) if conf_stats["max_conf"] else 0.0,
            "avg": round(float(conf_stats["avg_conf"]), 1) if conf_stats["avg_conf"] else 0.0,
            "p25": round(p25, 1),
            "p50": round(p50, 1),
            "p75": round(p75, 1),
        },
    }


@router.get("/jobs/{job_id}/report/mapping")
def parser_report_mapping(job_id: int, current_user: dict = Depends(require_auth())):
    """
    映射报告：从 parse_row_staging.source_columns JSON 聚合所有行的映射关系，
    找出每个标准字段对应哪些源列索引，以及一致性（consistency）。
    """
    _job_exists(job_id)

    TARGET_FIELDS = ["brand", "category", "model", "model_std", "price", "description", "notes"]

    with get_db() as db:
        with db.cursor() as cursor:
            cursor.execute("""
                SELECT source_columns
                FROM parse_row_staging
                WHERE job_id=%s AND source_columns IS NOT NULL AND source_columns != ''
            """, (job_id,))
            rows = cursor.fetchall()

            # Aggregate column indices per field
            field_collections = {f: [] for f in TARGET_FIELDS}
            for row in rows:
                try:
                    sc = row["source_columns"]
                    if isinstance(sc, str):
                        sc = json.loads(sc)
                    for f in TARGET_FIELDS:
                        if f in sc and sc[f] is not None:
                            field_collections[f].append(sc[f])
                except Exception:
                    continue

            cursor.execute("SELECT COUNT(*) as total FROM parse_row_staging WHERE job_id=%s", (job_id,))
            total = cursor.fetchone()["total"]

            field_mappings = []
            for f in TARGET_FIELDS:
                cols = field_collections[f]
                if cols:
                    unique_cols = list(set(cols))
                    # Consistency = most common column count / total rows with this field mapped
                    from collections import Counter
                    col_counts = Counter(cols)
                    most_common_count = col_counts.most_common(1)[0][1] if col_counts else 0
                    consistency = round(most_common_count / len(rows), 4) if rows else 0.0
                    field_mappings.append({
                        "field": f,
                        "source_columns": sorted(unique_cols),
                        "consistency": consistency,
                    })
                else:
                    field_mappings.append({
                        "field": f,
                        "source_columns": [],
                        "consistency": 0.0,
                    })

    return {
        "job_id": job_id,
        "field_mappings": field_mappings,
        "total_rows": total,
    }


@router.get("/jobs/{job_id}/report/standardization")
def parser_report_standardization(job_id: int, current_user: dict = Depends(require_auth())):
    """
    标准化报告：统计各字段填充率和 confidence 分位数。
    """
    _job_exists(job_id)

    TARGET_FIELDS = ["brand", "category", "model", "model_std", "price", "description", "notes"]

    with get_db() as db:
        with db.cursor() as cursor:
            cursor.execute("SELECT COUNT(*) as total FROM parse_row_staging WHERE job_id=%s", (job_id,))
            total = cursor.fetchone()["total"]

            field_completion = {}
            for f in TARGET_FIELDS:
                cursor.execute(f"""
                    SELECT COUNT(*) as cnt
                    FROM parse_row_staging
                    WHERE job_id=%s AND {f} IS NOT NULL AND {f} != ''
                """, (job_id,))
                field_completion[f] = cursor.fetchone()["cnt"]

            # Confidence stats
            cursor.execute("""
                SELECT confidence
                FROM parse_row_staging
                WHERE job_id=%s
                ORDER BY confidence ASC
            """, (job_id,))
            all_confs = [float(row["confidence"]) for row in cursor.fetchall()]
            n = len(all_confs)

            def percentile(vals, p):
                if not vals:
                    return 0.0
                idx = int((p / 100.0) * (len(vals) - 1))
                return vals[idx]

            if all_confs:
                conf_min = min(all_confs)
                conf_max = max(all_confs)
                conf_avg = round(sum(all_confs) / n, 1)
                conf_p25 = round(percentile(all_confs, 25), 1)
                conf_p75 = round(percentile(all_confs, 75), 1)
            else:
                conf_min = conf_max = conf_avg = conf_p25 = conf_p75 = 0.0

    return {
        "job_id": job_id,
        "field_completion": field_completion,
        "confidence": {
            "min": conf_min,
            "max": conf_max,
            "avg": conf_avg,
            "p25": conf_p25,
            "p75": conf_p75,
        },
    }


# ─── Export job results as Excel ─────────────────────────────────────────────────
@router.get("/jobs/{job_id}/export")
def parser_export_excel(job_id: int, current_user: dict = Depends(require_auth())):
    """Export parse job results as an Excel file."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    from urllib.parse import quote

    with get_db() as db:
        with db.cursor() as cursor:
            cursor.execute("SELECT job_code, original_filename, supplier_id FROM parse_jobs WHERE id=%s", (job_id,))
            job = cursor.fetchone()
            if not job:
                raise HTTPException(404, "任务不存在")

            cursor.execute("""
                SELECT sheet_name, row_index, row_status, brand, model, category, 
                       price, confidence, description, notes
                FROM parse_row_staging 
                WHERE job_id=%s 
                ORDER BY sheet_name, row_index
            """, (job_id,))
            rows = cursor.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "解析结果"

    headers_list = ["子表", "行号", "状态", "品牌", "型号", "品类", "价格", "置信度"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0'),
    )

    for col, h in enumerate(headers_list, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    valid_fill = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
    flagged_fill = PatternFill(start_color="FEFCE8", end_color="FEFCE8", fill_type="solid")

    for i, r in enumerate(rows, 2):
        status_text = "有效" if r["row_status"] == "valid" else "待审" if r["row_status"] == "flagged" else r["row_status"]
        price_val = float(r["price"]) if r["price"] is not None else None
        ws.cell(row=i, column=1, value=r["sheet_name"]).border = thin_border
        ws.cell(row=i, column=2, value=r["row_index"]).border = thin_border
        ws.cell(row=i, column=3, value=status_text).border = thin_border
        ws.cell(row=i, column=4, value=r["brand"] or "").border = thin_border
        ws.cell(row=i, column=5, value=r["model"] or "").border = thin_border
        ws.cell(row=i, column=6, value=r["category"] or "").border = thin_border
        ws.cell(row=i, column=7, value=price_val).border = thin_border
        ws.cell(row=i, column=8, value=float(r["confidence"])).border = thin_border
        fill = valid_fill if r["row_status"] == "valid" else flagged_fill
        for col in range(1, 9):
            ws.cell(row=i, column=col).fill = fill
            ws.cell(row=i, column=col).alignment = Alignment(horizontal='center')

    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 10

    ws2 = wb.create_sheet("汇总")
    total = len(rows)
    valid_cnt = sum(1 for r in rows if r["row_status"] == "valid")
    summary_data = [
        ("文件名", job["original_filename"]),
        ("任务编号", job["job_code"]),
        ("总行数", total),
        ("有效", valid_cnt),
        ("待审", sum(1 for r in rows if r["row_status"] == "flagged")),
        ("有效占比", f"{valid_cnt/total*100:.1f}%" if total else "0%"),
    ]
    for row_idx, (label, val) in enumerate(summary_data, 1):
        ws2.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws2.cell(row=row_idx, column=2, value=val)
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 30

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"{job['job_code']}_result.xlsx"
    encoded = quote(filename)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"}
    )


# ---------------------------------------------------------------------------
# 10. LLM consistency check — replaces heuristic confidence with DeepSeek
# ---------------------------------------------------------------------------

# In-memory task status store for async consistency checks
_consistency_tasks: dict[int, dict] = {}


@router.post("/jobs/{job_id}/check-consistency")
def check_consistency(
    job_id: int,
    background_tasks: BackgroundTasks,
    current_user: dict = Depends(require_auth()),
):
    """
    Kick off async LLM consistency check in background.
    Returns immediately with task status. Poll GET /consistency-status for result.
    """
    from parser_core import llm_check_consistency

    with get_db() as db:
        with db.cursor() as cursor:
            cursor.execute(
                "SELECT COUNT(*) as cnt FROM parse_row_staging WHERE job_id = %s",
                (job_id,),
            )
            row = cursor.fetchone()
            if not row or row["cnt"] == 0:
                raise HTTPException(404, "该任务没有数据行")

    _consistency_tasks[job_id] = {
        "status": "queued",
        "total": row["cnt"],
        "consistent": 0,
        "inconsistent": 0,
        "message": "等待处理",
    }

    background_tasks.add_task(_run_consistency_check, job_id)

    return {
        "job_id": job_id,
        "status": "queued",
        "total": row["cnt"],
        "message": "一致性检查已排队，请稍后查看结果",
    }


@router.get("/jobs/{job_id}/consistency-status")
def get_consistency_status(
    job_id: int,
    current_user: dict = Depends(require_auth()),
):
    """Poll this endpoint to check async consistency check progress."""
    task = _consistency_tasks.get(job_id)
    if not task:
        return {"job_id": job_id, "status": "not_started"}
    return {"job_id": job_id, **task}


def _run_consistency_check(job_id: int):
    """Background task runner for consistency check."""
    from parser_core import llm_check_consistency
    import json

    _consistency_tasks[job_id]["status"] = "running"

    with get_db() as db:
        with db.cursor() as cursor:
            cursor.execute(
                "SELECT id, raw_data, brand, model, price, category, manually_corrected "
                "FROM parse_row_staging WHERE job_id = %s "
                "ORDER BY sheet_name, row_index",
                (job_id,),
            )
            rows = cursor.fetchall()

    if not rows:
        _consistency_tasks[job_id]["status"] = "error"
        _consistency_tasks[job_id]["message"] = "无数据行"
        return

    batch_input = []
    for r in rows:
        raw = r.get("raw_data")
        if raw is None:
            raw = ""
        elif isinstance(raw, (dict, list)):
            raw = json.dumps(raw, ensure_ascii=False)
        batch_input.append({
            "raw_data": str(raw),
            "brand": r.get("brand") or "",
            "model": r.get("model") or "",
            "price": r.get("price"),
            "category": r.get("category") or "",
        })

    try:
        judgments = llm_check_consistency(batch_input)
    except Exception as e:
        _consistency_tasks[job_id]["status"] = "error"
        _consistency_tasks[job_id]["message"] = str(e)
        return

    consistent_count = 0
    inconsistent_count = 0
    error_count = 0
    skip_count = 0
    with get_db() as db:
        with db.cursor() as cursor:
            for j in judgments:
                idx = j["row_idx"]
                if idx < 0 or idx >= len(rows):
                    continue
                row = rows[idx]
                row_id = row["id"]
                is_consistent = j["consistent"]
                reason = j.get("reason", "")

                # Skip rows that were manually corrected
                if row.get("manually_corrected"):
                    skip_count += 1
                    continue

                if is_consistent is None:
                    # LLM error: keep original confidence, record error
                    error_count += 1
                    details = json.dumps({
                        "method": "llm_consistency",
                        "consistent": None,
                        "reason": reason,
                        "model": "deepseek-v4-flash",
                        "error": True,
                    })
                    cursor.execute(
                        "UPDATE parse_row_staging "
                        "SET confidence_details = %s, updated_at = NOW() "
                        "WHERE id = %s",
                        (details, row_id),
                    )
                elif is_consistent:
                    confidence = 95.0
                    consistent_count += 1
                    details = json.dumps({
                        "method": "llm_consistency",
                        "consistent": True,
                        "reason": reason,
                        "model": "deepseek-v4-flash",
                    })
                    cursor.execute(
                        "UPDATE parse_row_staging "
                        "SET confidence = %s, confidence_details = %s, updated_at = NOW() "
                        "WHERE id = %s",
                        (confidence, details, row_id),
                    )
                else:
                    confidence = 20.0
                    inconsistent_count += 1
                    details = json.dumps({
                        "method": "llm_consistency",
                        "consistent": False,
                        "reason": reason,
                        "model": "deepseek-v4-flash",
                    })
                    cursor.execute(
                        "UPDATE parse_row_staging "
                        "SET confidence = %s, confidence_details = %s, updated_at = NOW() "
                        "WHERE id = %s",
                        (confidence, details, row_id),
                    )
        db.commit()

    _consistency_tasks[job_id]["status"] = "done"
    _consistency_tasks[job_id]["consistent"] = consistent_count
    _consistency_tasks[job_id]["inconsistent"] = inconsistent_count
    _consistency_tasks[job_id]["error"] = error_count
    _consistency_tasks[job_id]["skipped"] = skip_count
    parts = []
    if consistent_count:
        parts.append(f"{consistent_count} 条一致")
    if inconsistent_count:
        parts.append(f"{inconsistent_count} 条不一致")
    if error_count:
        parts.append(f"{error_count} 条检查异常")
    if skip_count:
        parts.append(f"{skip_count} 条人工修正已跳过")
    _consistency_tasks[job_id]["message"] = (
        f"一致性检查完成：{'，'.join(parts)}" if parts else "无数据被更新"
    )

